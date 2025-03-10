from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import InstitutionRegisterForm, LoginForm, ClassHeadLoginForm, SubjectHeadLoginForm, StudentLoginForm, ParentLoginForm, AddClassForm, AddSubjectForm, AddStudentForm, ClassUploadForm, SubjectUploadForm, StudentUploadForm
from .models import Institution, Classes, Subjects, Students, Users, Parents, Chat, Announcements, Marks, Attendance, StudentEvaluation, QuizResponse
from django.db import IntegrityError, transaction
from django.db import connection
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
import json
from django.utils.timezone import now
from django.http import HttpResponse
from django.core.mail import send_mail
import pandas as pd
from django.core.files.storage import FileSystemStorage
import os, datetime
import openpyxl
from decimal import Decimal, ROUND_HALF_UP
import random
import re
from textblob import TextBlob

# Index page
def index(request):
    return render(request, 'index.html')


def send_account_creation_email(email, password, role, name, institution_email):
    subject = "Your Eduke Account is Ready"

    role_messages = {
        "class_head": "You have been assigned as a Class Head. You can manage students and subjects for your class.",
        "subject_head": "You have been assigned as a Subject Head. You can manage subjects and interact with students.",
        "student": "Your student account has been created. You can now access study materials and interact with teachers.",
        "parent": "Your parent account has been created. You can track your child's academic progress and communicate with teachers."
    }

    message = f"""
    Dear {name},

    Your Eduke account has been successfully created.

    Role: {role.replace('_', ' ').title()}
    Email: {email}
    Temporary Password: {password}

    {role_messages.get(role, "You can now access the system.")}

    Please log in and change your password.

    Regards,  
    Eduke Administration  
    ({institution_email})
    """

    send_mail(subject, message, institution_email, [email])



######################################################################################################################


# Institution Registration
def register_institution(request):
    if request.method == 'POST':
        form = InstitutionRegisterForm(request.POST)
        if form.is_valid():
            institution_name = form.cleaned_data['institution_name']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            # Check if the email already exists
            if Institution.objects.filter(email=email).exists():
                form.add_error('email', 'An institution with this email already exists.')
            else:
                try:
                    # Create the institution
                    Institution.objects.create(
                        institution_name=institution_name,
                        email=email,
                        password=password  # Store plain password here
                    )
                    messages.success(request, 'Institution registered successfully! Please log in.')
                    return redirect('login')
                except Exception as e:
                    form.add_error(None, 'An error occurred while saving the data. Please try again.')
    else:
        form = InstitutionRegisterForm()

    return render(request, 'registration/institution_register.html', {'form': form})

# Institution Login
def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            # Authenticate the institution  
            try:
                institution = Institution.objects.get(email=email, password=password)

                # Set session variables
                request.session['institution_id'] = institution.institution_id
                request.session['institution_name'] = institution.institution_name

                # Redirect to the admin dashboard
                return redirect('admin_dashboard')  # Adjust the URL name accordingly
            except Institution.DoesNotExist:
                messages.error(request, 'Invalid email or password.')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = LoginForm()

    return render(request, 'login.html', {'form': form})

# Admin Logout
def logout(request):
    # Clear session and logout
    request.session.flush()
    return redirect('/')

######################################################################################################################

# Admin Dashboard view
def admin_dashboard(request):
    # Check if the user is logged in
    if 'institution_id' not in request.session:
        messages.error(request, 'Please login to access the dashboard')
        return redirect('login')

    # Get the institution_id from the session
    institution_id = request.session['institution_id']
    
    try:
        # Fetch the institution based on the institution_id from the session
        institution = Institution.objects.get(institution_id=institution_id)

        # Current institution details
        institution_name = institution.institution_name
        password = institution.password
        email = institution.email
    except Institution.DoesNotExist:
        messages.error(request, "Institution not found.")
        return redirect('login')
    
    # Calculate the total number of users (class heads + students)
    total_users = Users.objects.filter(
        id__in=Classes.objects.filter(institution_id=institution_id).values('user')
    ).count() + Users.objects.filter(
        id__in=Students.objects.filter(class_obj__in=Classes.objects.filter(institution_id=institution_id).values('id')).values('user')
    ).count()

    # Get all classes belonging to the logged-in institution
    total_classes = Classes.objects.filter(institution_id=institution_id).count()

    # Count the students belonging to those classes
    student_count = Students.objects.filter(
        class_obj__in=Classes.objects.filter(institution_id=institution_id).values('id')
    ).count()

    # Count the subjects belonging to the logged-in institution (using class_obj field)
    subject_count = Subjects.objects.filter(
        class_obj__in=Classes.objects.filter(institution_id=institution_id).values('id')
    ).count()

    # Pass the counts to the template
    context = {
        'institution_name': institution_name,
        'total_users': total_users,
        'student_count': student_count,
        'total_classes': total_classes,
        'subject_count': subject_count,  # Add subject count to the context
    }

    return render(request, 'admin/admin_dashboard.html', context)


def admin_classes(request):
    print("🟢 View Loaded: admin_classes()")  

    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        print("🔴 User not logged in. Redirecting to login.")
        return redirect('login')

    institution_id = request.session['institution_id']
    print(f"🟢 Institution ID: {institution_id}")

    try:
        institution = Institution.objects.get(institution_id=institution_id)
        print(f"🟢 Institution Found: {institution.institution_name}")
    except Institution.DoesNotExist:
        messages.error(request, "Institution not found.")
        print("🔴 Institution not found. Redirecting to login.")
        return redirect('login')

    classes = Classes.objects.filter(institution_id=institution.institution_id)
    print(f"🟢 Retrieved {classes.count()} classes.")
    for class_instance in classes:
        print(f"   Class ID: {class_instance.id}, Name: {class_instance.class_name}, Head: {class_instance.class_head}, Email: {class_instance.email}")

    if request.method == 'POST':
        print("🟢 POST Request Received.")
        print(f"🔍 Request Data: {request.POST}")

        print("🟢 Handling New Class Addition")
        form = AddClassForm(request.POST)

        if form.is_valid():
            with transaction.atomic():
                class_name = form.cleaned_data['class_name']
                class_head = form.cleaned_data['class_head']
                email = form.cleaned_data['email']
                password = form.cleaned_data['password']

                print(f"🟢 New Class Data - Name: {class_name}, Head: {class_head}, Email: {email}")

                user = Users.objects.create(role='class_head')
                print("🟢 New User Created with ID:", user.id)

                new_class = Classes(
                    class_name=class_name,
                    class_head=class_head,
                    email=email,
                    password=password,
                    institution=institution,
                    user=user
                )
                new_class.save()
                print(f"🟢 New Class Added with ID: {new_class.id}. Data: Name: {class_name}, Head: {class_head}, Email: {email}")

                send_account_creation_email(email, password, "class_head", class_head, institution.email)
                print("🟢 Account creation email sent.")

                messages.success(request, 'New class added successfully!')
                return redirect('admin_classes')
        else:
            messages.error(request, 'Error adding class. Please try again.')
            print("🔴 Error: Form validation failed.")

    else:
        print("🟢 Rendering Page with GET Request.")
        form = AddClassForm()

    context = {
        'institution_name': institution.institution_name,
        'classes': classes,
        'form': form,
    }

    print(f"🟢 Sending context to template: {context}")
    print("🟢 Rendering admin/admin_classes.html")
    return render(request, 'admin/admin_classes.html', context)


def admin_class_edit(request, class_id):
    try:
        # Fetch the class object to be edited
        class_obj = Classes.objects.get(id=class_id)

        if request.method == "POST":
            # Update class details based on the form input
            class_obj.class_name = request.POST.get('class_name')
            class_obj.class_head = request.POST.get('class_head')
            class_obj.email = request.POST.get('email')
            class_obj.save()

            # Redirect to the class list page after successful update
            return redirect('admin_classes')  # Adjust to the correct URL pattern for the class list

        # Render the edit page with the class details
        return render(request, 'admin/class_edit.html', {'class': class_obj})
    
    except Classes.DoesNotExist:
        # Handle the case when the class doesn't exist
        return redirect('admin_classes')  # Redirect to the class list if class not found


def admin_class_detail(request, class_id):
    # Check if the user is logged in
    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        return redirect('login')

    # Get the institution_id from the session
    institution_id = request.session['institution_id']

    with connection.cursor() as cursor:
        # Fetch institution details
        cursor.execute("SELECT institution_name FROM main_institution WHERE institution_id = %s", [institution_id])
        institution_row = cursor.fetchone()
        if not institution_row:
            messages.error(request, "Institution not found.")
            return redirect('login')
        institution_name = institution_row[0]

        # Fetch class details
        cursor.execute("SELECT id, class_name, class_head, email, password, user_id FROM main_classes WHERE id = %s", [class_id])
        class_row = cursor.fetchone()
        if not class_row:
            messages.error(request, "Class not found.")
            return redirect('admin_students')

        class_obj = {
            'id': class_row[0],
            'class_name': class_row[1],
            'class_head': class_row[2],
            'email': class_row[3],
            'password': class_row[4],
            'user_id': class_row[5],
        }

        # Fetch students belonging to the class
        cursor.execute("SELECT id, roll_no, email, name FROM main_students WHERE class_obj_id = %s", [class_id])
        students = [{'id': row[0], 'roll_no': row[1], 'name': row[2], 'email': row[3]} for row in cursor.fetchall()]

    # Prepare context for the template
    context = {
        'institution_name': institution_name,
        'class_obj': class_obj,
        'students': students,
        'class_head': class_obj['class_head'],  # Added class_head to the context
        'class_email': class_obj['email'],     # Added class email to the context
    }

    return render(request, 'admin/admin_class_detail.html', context)



def admin_subjects(request):
    # Check if the user is logged in
    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        return redirect('login')
    
    # Get the institution_id from the session
    institution_id = request.session['institution_id']
    print(f"DEBUG: Institution ID from session - {institution_id}")

    try:
        # Fetch the institution based on the session institution_id
        institution = Institution.objects.get(institution_id=institution_id)
        print(f"DEBUG: Institution found - {institution.institution_name}")
    except Institution.DoesNotExist:
        messages.error(request, "Institution not found.")
        return redirect('login')

    # Fetch all classes for the dropdown
    classes = Classes.objects.filter(institution_id=institution_id).values('id', 'class_name')
    print(f"DEBUG: Classes retrieved - {list(classes)}")

    # Fetch subjects along with class names
    subjects = []
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                ms.id, ms.subject_name, ms.subject_head, ms.email, 
                c.class_name
            FROM main_subjects ms
            JOIN main_classes c ON ms.class_obj_id = c.id
            WHERE c.institution_id = %s
        """, [institution_id])
        subjects = cursor.fetchall()

    print(f"DEBUG: Retrieved subjects - {subjects}")

    if request.method == "POST":
        form = AddSubjectForm(request.POST)
        if form.is_valid():
            subject_name = form.cleaned_data['subject_name']
            subject_head = form.cleaned_data['subject_head']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            class_id = form.cleaned_data['class_id'].id  # Get class ID

            print(f"DEBUG: Form Data - Subject: {subject_name}, Head: {subject_head}, Email: {email}, Class ID: {class_id}")

            with connection.cursor() as cursor:
                try:
                    # Step 1: Insert into main_users table
                    cursor.execute("INSERT INTO main_users (role) VALUES (%s)", ['subject_head'])
                    user_id = cursor.lastrowid  # Get the last inserted user ID
                    
                    if not user_id:
                        raise Exception("User ID not retrieved after insert!")

                    print(f"DEBUG: Inserted into main_users - User ID: {user_id}")

                    # Step 2: Insert into main_subjects table
                    cursor.execute("""
                        INSERT INTO main_subjects (subject_name, subject_head, email, password, user_id, class_obj_id) 
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, [subject_name, subject_head, email, password, user_id, class_id])

                    print(f"DEBUG: Inserted into main_subjects - Subject: {subject_name}, Class ID: {class_id}")

                    # **Step 3: Send Account Creation Email**
                    send_account_creation_email(email, password, "subject_head", subject_head, institution.email)

                    messages.success(request, "Subject added successfully!")
                    return redirect('admin_subjects')

                except Exception as e:
                    print(f"ERROR: {e}")
                    messages.error(request, f"An error occurred: {e}")
        else:
            print("DEBUG: Form is invalid")

    else:
        form = AddSubjectForm()

    context = {
        'institution_name': institution.institution_name,
        'form': form,
        'classes': classes,
        'subjects': subjects  # Pass subjects data to template
    }

    return render(request, 'admin/admin_subjects.html', context)



def admin_subject_edit(request, subject_id):
    try:
        # Fetch the subject object to be edited
        subject_obj = Subjects.objects.get(id=subject_id)

        if request.method == "POST":
            # Update subject details based on the form input
            subject_obj.subject_name = request.POST.get('subject_name')
            subject_obj.subject_head = request.POST.get('subject_head')
            subject_obj.email = request.POST.get('email')
            subject_obj.class_id = request.POST.get('class_id')
            subject_obj.save()

            # Redirect to the subject list page after successful update
            return redirect('admin_subjects')  # Adjust to the correct URL pattern for the subject list

        # Render the edit page with the subject details
        return render(request, 'admin/subject_edit.html', {'subject': subject_obj})
    
    except Subjects.DoesNotExist:
        # Handle the case when the subject doesn't exist
        return redirect('admin_subjects')  # Redirect to the subject list if subject not found



def admin_subject_detail(request, subject_id):
    print(f"Received subject_id: {subject_id}, Type: {type(subject_id)}")

    # Validate subject_id
    if not subject_id or not str(subject_id).isdigit():
        print("Invalid subject_id, redirecting to admin_students.")
        messages.error(request, "Invalid subject ID.")
        return redirect('admin_students')

    # Check if the user is logged in
    if 'institution_id' not in request.session:
        print("User not logged in, redirecting to login page.")
        messages.error(request, 'Please log in to access this page.')
        return redirect('login')

    # Get the institution_id from the session
    institution_id = request.session['institution_id']
    print(f"Institution ID from session: {institution_id}")

    with connection.cursor() as cursor:
        # Fetch institution details
        cursor.execute("SELECT institution_name FROM main_institution WHERE institution_id = %s", [institution_id])
        institution_row = cursor.fetchone()
        if not institution_row:
            print("Institution not found, redirecting to login.")
            messages.error(request, "Institution not found.")
            return redirect('login')
        institution_name = institution_row[0]
        print(f"Institution Name: {institution_name}")

        # Fetch subject details
        cursor.execute("SELECT id, subject_name, subject_head, email FROM main_subjects WHERE id = %s", [subject_id])
        subject_row = cursor.fetchone()
        if not subject_row:
            print("Subject not found, redirecting to admin_students.")
            messages.error(request, "Subject not found.")
            return redirect('admin_students')

        subject_obj = {
            'id': subject_row[0],
            'subject_name': subject_row[1],
            'subject_head': subject_row[2],
            'email': subject_row[3],
        }
        print(f"Subject Name: {subject_obj['subject_name']}")

    # Prepare context for the template
    context = {
        'institution_name': institution_name,
        'subject_obj': subject_obj,
    }

    return render(request, 'admin/admin_subject_detail.html', context)



def admin_students(request):
    # Check if the user is logged in
    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        return redirect('login')

    # Get the institution_id from the session
    institution_id = request.session['institution_id']
    print(f"DEBUG: Institution ID from session - {institution_id}")

    try:
        # Fetch the institution based on the session institution_id
        institution = Institution.objects.get(institution_id=institution_id)
        print(f"DEBUG: Institution found - {institution.institution_name}")
    except Institution.DoesNotExist:
        messages.error(request, "Institution not found.")
        return redirect('login')

    # Fetch all classes for the dropdown
    classes = Classes.objects.filter(institution_id=institution_id)
    print(f"DEBUG: Classes retrieved - {list(classes.values('id', 'class_name'))}")

    # Fetch students along with class name and class head name
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                s.id, s.name AS student_name, s.roll_no, s.email, c.class_name, 
                c.class_head
            FROM main_students s
            LEFT JOIN main_classes c ON s.class_obj_id = c.id
            WHERE c.institution_id = %s
        """, [institution_id])
        students = cursor.fetchall()

    print(f"DEBUG: Students retrieved - {students}")

    # Handle form submission
    if request.method == "POST":
        form = AddStudentForm(request.POST)

        if form.is_valid():
            name = form.cleaned_data['name']
            roll_no = int(form.cleaned_data['roll_no'])  # Ensure integer conversion
            email = form.cleaned_data['email']  # Extract email
            password = form.cleaned_data['password']
            class_obj_id = form.cleaned_data['class_id'].id  # Extracting the integer ID

            print(f"DEBUG: Form Data - Name: {name}, Roll No: {roll_no}, Email: {email}, Password: {password}, Class ID: {class_obj_id}")

            try:
                with transaction.atomic():
                    with connection.cursor() as cursor:
                        # Insert into users table (Student)
                        cursor.execute("INSERT INTO main_users (role) VALUES ('student')")
                        student_user_id = cursor.lastrowid
                        print(f"DEBUG: Inserted into users table (student), User ID: {student_user_id}")

                        # Insert student record (Including Email)
                        cursor.execute("""
                            INSERT INTO main_students (name, roll_no, email, password, class_obj_id, user_id)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, [name, roll_no, email, password, class_obj_id, student_user_id])
                        student_id = cursor.lastrowid
                        print(f"DEBUG: Inserted into students table, Student ID: {student_id}")

                        # Insert into users table (Parent)
                        cursor.execute("INSERT INTO main_users (role) VALUES ('parent')")
                        parent_user_id = cursor.lastrowid
                        print(f"DEBUG: Inserted into users table (parent), User ID: {parent_user_id}")

                        # Insert parent record
                        cursor.execute("""
                            INSERT INTO main_parents (student_id, password, name, user_id)
                            VALUES (%s, %s, NULL, %s)
                        """, [student_id, password, parent_user_id])
                        print(f"DEBUG: Inserted into parents table, Parent User ID: {parent_user_id}")

                    # Send account creation email to student
                    send_account_creation_email(email, password, "student", name, institution.email)

                    messages.success(request, "Student and Parent added successfully!")
                    return redirect('admin_students')

            except Exception as e:
                print(f"ERROR: {e}")
                messages.error(request, f"Error: {str(e)}")
        else:
            print("DEBUG: Form validation failed")
            messages.error(request, "Failed to add student. Please check the form.")
    else:
        form = AddStudentForm()

    context = {
        'institution_name': institution.institution_name,
        'form': form,
        'students': students,
        'classes': classes,
    }

    return render(request, 'admin/admin_students.html', context)


def admin_student_edit(request, student_id):
    try:
        # Fetch the student object to be edited
        student_obj = Students.objects.get(id=student_id)

        if request.method == "POST":
            # Update student details based on the form input
            student_obj.name = request.POST.get('name')
            student_obj.email = request.POST.get('email')
            student_obj.roll_no = request.POST.get('roll_no')
            student_obj.class_name = request.POST.get('class')
            student_obj.save()

            # Redirect to the student list page after successful update
            return redirect('admin_students')  # Adjust to the correct URL pattern for the student list

        # Render the edit page with the student details
        return render(request, 'admin/student_edit.html', {'student': student_obj})
    
    except Students.DoesNotExist:
        # Handle the case when the student doesn't exist
        return redirect('admin_students')  # Redirect to the student list if student not found



from django.shortcuts import render, redirect
from django.db import connection
from django.contrib import messages

def admin_student_performance(request, student_id):
    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        return redirect('login')
    
    institution_id = request.session['institution_id']
    print(f"DEBUG: Institution ID from session - {institution_id}")
    
    print(f"Received student_id: {student_id}, Type: {type(student_id)}")

    # Validate student_id
    if not student_id or not str(student_id).isdigit():
        print("Invalid student_id, redirecting to admin_students.")
        messages.error(request, "Invalid student ID.")
        return redirect('admin_students')

    try:
        with connection.cursor() as cursor:
            # Fetch student details (excluding parent details)
            cursor.execute("""
                SELECT 
                    s.id, s.roll_no, s.name, s.email, s.class_obj_id, c.class_name
                FROM main_students s
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s
            """, [student_id])
            student_row = cursor.fetchone()
            
            if not student_row:
                print("Student not found, redirecting to admin_students.")
                messages.error(request, "Student not found.")
                return redirect('admin_students')

            student = {
                'id': student_row[0],
                'roll_no': student_row[1],
                'name': student_row[2],
                'email': student_row[3],
                'class_id': student_row[4],
                'class_name': student_row[5],
            }
            print(f"Student Name: {student['name']}")

        # Extract subject_id from GET request
        subject_id = request.GET.get('subject_id', None)

        # Fetch subjects for the student's class
        subjects = []
        if student.get("class_id"):
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, subject_name FROM main_subjects WHERE class_obj_id = %s
                """, [student["class_id"]])
                subjects = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]

        # Get the name of the selected subject
        selected_subject_name = None
        if subject_id:
            selected_subject_name = next((s["name"] for s in subjects if str(s["id"]) == subject_id), None)
            print(f"[DEBUG] Selected Subject: {selected_subject_name} (ID: {subject_id})")

        # Initialize default evaluation data
        evaluations = {
            "marks_percentage": 0,
            "attendance_percentage": 0,
            "study_time_rating": 0,
            "sleep_time_rating": 0,
            "class_participation_rating": 0,
            "academic_activity_rating": 0,
        }

        # Fetch student evaluation details for the selected subject
        if subject_id:
            print(f"[DEBUG] Fetching evaluation data for subject: {selected_subject_name} (ID: {subject_id})...")
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT marks_percentage, attendance_percentage,
                           study_time_rating, sleep_time_rating, 
                           class_participation_rating, academic_activity_rating
                    FROM main_studentevaluation
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    evaluations = {
                        "marks_percentage": round(row[0] or 0, 2),
                        "attendance_percentage": round(row[1] or 0, 2),
                        "study_time_rating": round(row[2] or 0, 2),
                        "sleep_time_rating": round(row[3] or 0, 2),
                        "class_participation_rating": round(row[4] or 0, 2),
                        "academic_activity_rating": round(row[5] or 0, 2),
                    }
                    print(f"[DEBUG] Evaluation Data: {evaluations}")
                else:
                    print(f"[WARNING] No evaluation data found for {selected_subject_name}.")

        # Fetch quiz performance for the selected subject
        quiz_percentage = 0
        if subject_id:
            print(f"[DEBUG] Fetching quiz performance for {selected_subject_name} (ID: {subject_id})...")
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT COUNT(*) AS total_attempted,
                           SUM(CASE WHEN q.correct_option = r.student_response THEN 1 ELSE 0 END) AS correct_answers
                    FROM main_quizresponse r
                    JOIN main_quizquestions q ON r.question_id = q.id
                    JOIN main_quizzes mq ON q.quiz_id = mq.id
                    WHERE r.student_id = %s AND mq.subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    total_attempted, correct_answers = row[0], row[1] or 0
                    quiz_percentage = round((correct_answers / total_attempted) * 100, 2) if total_attempted > 0 else 0
                    print(f"[DEBUG] Quiz Performance: Attempted = {total_attempted}, Correct = {correct_answers}, Percentage = {quiz_percentage}%")
                else:
                    print(f"[WARNING] No quiz data found for {selected_subject_name}.")

        # Prepare data for Graph
        graph_data = [
            float(evaluations.get("marks_percentage", 0)),
            float(evaluations.get("attendance_percentage", 0)),
            float(evaluations.get("study_time_rating", 0)),
            float(evaluations.get("sleep_time_rating", 0)),
            float(evaluations.get("class_participation_rating", 0)),
            float(evaluations.get("academic_activity_rating", 0)),
            float(quiz_percentage)
        ]

    except Exception as e:
        print(f"[ERROR] An error occurred: {e}")
        return redirect('error_page')  # Redirect to an error page if needed

    # Prepare context for the template
    context = {
        'student': student,
        'subjects': subjects,
        'selected_subject_id': subject_id,
        'evaluations': evaluations,
        'quiz_percentage': quiz_percentage,
        "graph_data": json.dumps(graph_data)
    }

    return render(request, 'admin/admin_student_performance.html', context)





def admin_profile(request):
    # Check if the user is logged in (session variable 'institution_id' is set)
    if 'institution_id' not in request.session:
        messages.error(request, "Please log in to access the admin profile.")
        return redirect('login')

    # Get the institution_id from the session
    institution_id = request.session['institution_id']
    
    try:
        # Fetch the institution based on the institution_id from the session
        institution = Institution.objects.get(institution_id=institution_id)

        # Current institution details
        institution_name = institution.institution_name
        password = institution.password
        email = institution.email
    except Institution.DoesNotExist:
        messages.error(request, "Institution not found.")
        return redirect('login')

    # Handle form submission (POST request)
    if request.method == 'POST':
        # Update the institution name and password if they are provided in the form
        new_institution_name = request.POST.get('institution_name')
        new_password = request.POST.get('password')

        # Update the institution details in the database
        if new_institution_name:
            institution.institution_name = new_institution_name
        if new_password:
            institution.password = new_password

        institution.save()  # Save the updated institution details

        messages.success(request, 'Your profile has been updated successfully.')
        return redirect('admin_profile')  # Redirect to prevent re-posting on refresh

    # Render the profile page with the current details
    context = {
        'institution_name': institution_name,
        'password': password,
        'email': email
    }
    
    return render(request, 'admin/admin_profile.html', context)


######################################################################################################################


def upload_classes(request):
    print("Request received. Method:", request.method)  # Debugging

    if request.method == 'POST':
        form = ClassUploadForm(request.POST, request.FILES)
        if form.is_valid():
            print("Form is valid.")  # Debugging
        else:
            print("Form errors:", form.errors)  # Debugging
            messages.add_message(request, messages.ERROR, "❌ Invalid form submission.", extra_tags="classes_error")
            return render(request, 'admin/admin_classes.html', {'form': form})

        file = request.FILES['file']
        print("File received:", file.name)  # Debugging

        # Save file temporarily
        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        file_path = fs.path(filename)
        print("File saved at:", file_path)  # Debugging

        try:
            # Fetch the currently logged-in institution's ID
            institution_id = request.session.get('institution_id')  # Assuming it's stored in session
            if not institution_id:
                messages.add_message(request, messages.ERROR, "❌ Institution not found. Please log in again.", extra_tags="classes_error")
                print("Error: Institution ID not found.")  # Debugging
                return redirect('admin_classes')

            print("Logged-in Institution ID:", institution_id)  # Debugging

            # Read Excel file
            try:
                df = pd.read_excel(file_path)
                print("File read successfully.")  # Debugging
                print("Columns in file:", df.columns.tolist())  # Debugging
            except Exception as e:
                messages.add_message(request, messages.ERROR, f"❌ Error reading the file: {e}", extra_tags="classes_error")
                print("Error reading the file:", e)  # Debugging
                return redirect('admin_classes')

            # Validate required columns
            required_columns = ['Class Name', 'Class Head', 'Email', 'Password']
            if not all(col in df.columns for col in required_columns):
                messages.add_message(request, messages.ERROR, "❌ Invalid file format! Ensure the correct columns are present.", extra_tags="classes_error")
                print("Error: Missing required columns.")  # Debugging
                return redirect('admin_classes')

            with connection.cursor() as cursor:
                for index, row in df.iterrows():
                    class_name = row['Class Name']
                    class_head = row['Class Head']
                    email = row['Email']
                    password = row['Password']

                    print(f"Processing row {index}: {class_name}, {class_head}, {email}")  # Debugging

                    try:
                        # Insert into Users table
                        cursor.execute("INSERT INTO main_users (role) VALUES (%s)", ['class_head'])
                        user_id = cursor.lastrowid  # Get the last inserted ID
                        print(f"Inserted user with ID: {user_id}")  # Debugging

                        # Insert into Classes table using the logged-in institution_id
                        cursor.execute("""
                            INSERT INTO main_classes (institution_id, class_name, class_head, email, password, user_id)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, [institution_id, class_name, class_head, email, password, user_id])
                        print(f"Inserted class: {class_name}")  # Debugging

                    except Exception as e:
                        messages.add_message(request, messages.ERROR, f"❌ Error inserting class {class_name}: {e}", extra_tags="classes_error")
                        print(f"Error inserting class {class_name}:", e)  # Debugging
                        continue  # Skip this row

            messages.add_message(request, messages.SUCCESS, "✅ Classes uploaded successfully!", extra_tags="classes_success")
            print("Upload successful! Redirecting to admin_classes.")  # Debugging
            return redirect('admin_classes')

        except Exception as e:
            messages.add_message(request, messages.ERROR, f"❌ Error processing file: {e}", extra_tags="classes_error")
            print("Error occurred:", e)  # Debugging
            return redirect('admin_classes')

    else:
        form = ClassUploadForm()
        print("GET request received. Rendering form.")  # Debugging

    return render(request, 'admin/admin_classes.html', {'form': form})



def upload_subjects(request):
    print("Request received. Method:", request.method)  # Debugging

    if request.method == 'POST':
        form = SubjectUploadForm(request.POST, request.FILES)
        if form.is_valid():
            print("Form is valid.")  # Debugging
        else:
            print("Form errors:", form.errors)  # Debugging
            messages.add_message(request, messages.ERROR, "❌ Invalid form submission.", extra_tags="subject_error")
            return render(request, 'admin/admin_subjects.html', {'form': form})

        file = request.FILES['file']
        print("File received:", file.name)  # Debugging

        # Save file temporarily
        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        file_path = fs.path(filename)
        print("File saved at:", file_path)  # Debugging

        try:
            # Fetch the currently logged-in institution's ID
            institution_id = request.session.get('institution_id')  # Assuming it's stored in session
            if not institution_id:
                messages.add_message(request, messages.ERROR, "❌ Institution not found. Please log in again.", extra_tags="subject_error")
                print("Error: Institution ID not found.")  # Debugging
                return redirect('admin_subjects')

            print("Logged-in Institution ID:", institution_id)  # Debugging

            # Read Excel file
            try:
                df = pd.read_excel(file_path)
                print("File read successfully.")  # Debugging
                print("Columns in file:", df.columns.tolist())  # Debugging
            except Exception as e:
                messages.add_message(request, messages.ERROR, f"❌ Error reading the file: {e}", extra_tags="subject_error")
                print("Error reading the file:", e)  # Debugging
                return redirect('admin_subjects')

            # Validate required columns
            required_columns = ['Subject Name', 'Subject Head', 'Email', 'Password', 'Class Name']
            if not all(col in df.columns for col in required_columns):
                messages.add_message(request, messages.ERROR, "❌ Invalid file format! Ensure the correct columns are present.", extra_tags="subject_error")
                print("Error: Missing required columns.")  # Debugging
                return redirect('admin_subjects')

            with connection.cursor() as cursor:
                for index, row in df.iterrows():
                    subject_name = row['Subject Name']
                    subject_head = row['Subject Head']
                    email = row['Email']
                    password = row['Password']
                    class_name = row['Class Name']

                    print(f"Processing row {index}: {subject_name}, {subject_head}, {email}, {class_name}")  # Debugging

                    # Fetch the class_obj_id based on Class Name
                    cursor.execute("SELECT id FROM main_classes WHERE class_name = %s AND institution_id = %s", [class_name, institution_id])
                    class_row = cursor.fetchone()

                    if class_row:
                        class_obj_id = class_row[0]
                        print(f"Class '{class_name}' found with ID: {class_obj_id}")  # Debugging
                    else:
                        messages.add_message(request, messages.ERROR, f"❌ Class '{class_name}' does not exist. Fix the Excel file.", extra_tags="subject_error")
                        print(f"Error: Class '{class_name}' not found.")  # Debugging
                        continue  # Skip this row

                    try:
                        # Insert into Users table
                        cursor.execute("INSERT INTO main_users (role) VALUES (%s)", ['subject_head'])
                        user_id = cursor.lastrowid  # Get the last inserted ID
                        print(f"Inserted user with ID: {user_id}")  # Debugging

                        # Insert into main_subjects table using class_obj_id
                        cursor.execute("""
                            INSERT INTO main_subjects (class_obj_id, subject_name, subject_head, email, password, user_id)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, [class_obj_id, subject_name, subject_head, email, password, user_id])
                        print(f"Inserted subject: {subject_name}")  # Debugging

                    except Exception as e:
                        messages.add_message(request, messages.ERROR, f"❌ Error inserting subject {subject_name}: {e}", extra_tags="subject_error")
                        print(f"Error inserting subject {subject_name}:", e)  # Debugging
                        continue  # Skip this row

            messages.add_message(request, messages.SUCCESS, "✅ Subjects uploaded successfully!", extra_tags="subject_success")
            print("Upload successful! Redirecting to admin_subjects.")  # Debugging
            return redirect('admin_subjects')

        except Exception as e:
            messages.add_message(request, messages.ERROR, f"❌ Error processing file: {e}", extra_tags="subject_error")
            print("Error occurred:", e)  # Debugging
            return redirect('admin_subjects')

    else:
        form = SubjectUploadForm()
        print("GET request received. Rendering form.")  # Debugging

    return render(request, 'admin/admin_subjects.html', {'form': form})



def upload_students(request):
    print("Request received. Method:", request.method)  # Debugging
    error_occurred = False  # Flag to track if errors occur

    if request.method == 'POST':
        form = StudentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            print("Form is valid.")  # Debugging
        else:
            print("Form errors:", form.errors)  # Debugging
            messages.add_message(request, messages.ERROR, "❌ Invalid form submission.", extra_tags="student_error")
            return render(request, 'admin/admin_students.html', {'form': form})

        if form.is_valid():
            file = request.FILES['file']
            print("File received:", file.name)  # Debugging

            # Save file temporarily
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            file_path = fs.path(filename)
            print("File saved at:", file_path)  # Debugging

            try:
                # Fetch the currently logged-in institution's ID
                institution_id = request.session.get('institution_id')  # Assuming it's stored in session
                if not institution_id:
                    messages.add_message(request, messages.ERROR, "❌ Institution not found. Please log in again.", extra_tags="student_error")
                    print("Error: Institution ID not found.")  # Debugging
                    return render(request, 'admin/admin_students.html', {'form': form})

                print("Logged-in Institution ID:", institution_id)  # Debugging

                # Read Excel file
                try:
                    df = pd.read_excel(file_path)
                    print("File read successfully.")  # Debugging
                    print("Columns in file:", df.columns.tolist())  # Debugging
                except Exception as e:
                    messages.add_message(request, messages.ERROR, f"❌ Error reading the file: {e}", extra_tags="student_error")
                    print("Error reading the file:", e)  # Debugging
                    return render(request, 'admin/admin_students.html', {'form': form})

                # Validate required columns
                required_columns = ['Student Name', 'Email', 'Roll No', 'Password', 'Class Name']
                if not all(col in df.columns for col in required_columns):
                    messages.add_message(request, messages.ERROR, "❌ Invalid file format! Ensure the correct columns are present.", extra_tags="student_error")
                    print("Error: Missing required columns.")  # Debugging
                    return render(request, 'admin/admin_students.html', {'form': form})

                with connection.cursor() as cursor:
                    for index, row in df.iterrows():
                        student_name = row['Student Name']
                        email = row['Email']
                        roll_no = row['Roll No']
                        password = row['Password']
                        class_name = row['Class Name']

                        print(f"Processing row {index}: {student_name}, {email}, {roll_no}, {class_name}")  # Debugging

                        # Fetch the class_obj_id based on Class Name
                        cursor.execute("SELECT id FROM main_classes WHERE class_name = %s AND institution_id = %s", [class_name, institution_id])
                        class_row = cursor.fetchone()

                        if class_row:
                            class_obj_id = class_row[0]
                            print(f"Class '{class_name}' found with ID: {class_obj_id}")  # Debugging
                        else:
                            messages.add_message(request, messages.ERROR, f"❌ Class '{class_name}' does not exist. Fix the Excel file.", extra_tags="student_error")
                            print(f"Error: Class '{class_name}' not found.")  # Debugging
                            error_occurred = True
                            continue  # Skip this row

                        try:
                            # Insert into Users table for student
                            cursor.execute("INSERT INTO main_users (role) VALUES (%s)", ['student'])
                            student_user_id = cursor.lastrowid  # Get last inserted ID
                            print(f"Inserted student user with ID: {student_user_id}")  # Debugging

                            # Insert into Students table
                            cursor.execute("""
                                INSERT INTO main_students (user_id, name, roll_no, password, class_obj_id, email)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            """, [student_user_id, student_name, roll_no, password, class_obj_id, email])
                            student_id = cursor.lastrowid  # Get last inserted student's ID
                            print(f"Inserted student: {student_name} with ID: {student_id}")  # Debugging

                            # Insert into Users table for parent
                            cursor.execute("INSERT INTO main_users (role) VALUES (%s)", ['parent'])
                            parent_user_id = cursor.lastrowid  # Get last inserted ID
                            print(f"Inserted parent user with ID: {parent_user_id}")  # Debugging

                            # Insert into Parents table using student_id
                            cursor.execute("""
                                INSERT INTO main_parents (user_id, student_id, password, name)
                                VALUES (%s, %s, %s, NULL)
                            """, [parent_user_id, student_id, roll_no])
                            print(f"Inserted parent for student: {student_name}")  # Debugging

                        except Exception as e:
                            messages.add_message(request, messages.ERROR, f"❌ Error inserting student {student_name}: {e}", extra_tags="student_error")
                            print(f"Error inserting student {student_name}:", e)  # Debugging
                            error_occurred = True
                            continue  # Skip this row

                if error_occurred:
                    messages.add_message(request, messages.WARNING, "⚠️ Some students could not be uploaded due to errors.", extra_tags="student_error")
                else:
                    messages.add_message(request, messages.SUCCESS, "✅ Students uploaded successfully!", extra_tags="student_success")

                print("Upload complete. Redirecting to admin_students.")  # Debugging
                return redirect('admin_students')

            except Exception as e:
                messages.add_message(request, messages.ERROR, f"❌ Error processing file: {e}", extra_tags="student_error")
                print("Error occurred:", e)  # Debugging
                return render(request, 'admin/admin_students.html', {'form': form})

    else:
        form = StudentUploadForm()
        print("GET request received. Rendering form.")  # Debugging

    return render(request, 'admin/admin_students.html', {'form': form})







######################################################################################################################


def class_head_login(request):
    form = ClassHeadLoginForm()

    if request.method == 'POST':
        form = ClassHeadLoginForm(request.POST)

        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            # Fetch class head record
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT c.id, c.password, u.role
                    FROM main_classes c
                    JOIN main_users u ON c.user_id = u.id
                    WHERE c.email = %s
                """, [email])

                class_head = cursor.fetchone()

            if class_head:
                class_id, db_password, role = class_head

                # Validate role and password
                if role == 'class_head' and db_password == password:
                    request.session['class_id'] = class_id  # Store class_id in session
                    return redirect('class_head_dashboard')  # Redirect to the dashboard
                else:
                    messages.error(request, 'Invalid credentials or access restricted.')
            else:
                messages.error(request, 'No Class Head found with this email.')

    return render(request, 'class_head/class_head_login.html', {'form': form})



def class_head_dashboard(request):
    class_id = request.session.get('class_id')

    # If no class_id in session, redirect to login
    if not class_id:
        messages.error(request, 'Session expired or not logged in. Please log in again.')
        return redirect('class_head_login')

    # Fetch class details for the logged-in class head
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT c.class_name, c.email, c.class_head
            FROM main_classes c
            WHERE c.id = %s
        """, [class_id])
        class_details = cursor.fetchone()

    # Ensure class_details is not None
    if not class_details:
        messages.error(request, 'Class details not found.')
        return redirect('class_head_login')

    # Fetch student count
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(*)
            FROM main_students s
            WHERE s.class_obj_id = %s
        """, [class_id])
        student_count = cursor.fetchone()[0]  # Fetch count of students in this class

    # Fetch subject count
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(*)
            FROM main_subjects sub
            WHERE sub.class_obj_id = %s
        """, [class_id])
        subject_count = cursor.fetchone()[0]  # Fetch count of subjects for this class

    return render(request, 'class_head/class_head_dashboard.html', {
        'class_details': class_details,
        'student_count': student_count,
        'subject_count': subject_count,  # Add subject count to template context
    })


def class_head_class(request):
    class_id = request.session.get('class_id')
    if not class_id:
        return redirect('class_head_login')  # Redirect to login if not logged in
    
    print(f"Current Class ID: {class_id}")

    try:
        class_details = Classes.objects.get(id=class_id)  # Retrieve class using the class_id
    except Classes.DoesNotExist:
        messages.error(request, 'Class not found.')
        return redirect('class_head_login')
    
    # Fetch the actual user_id from main_users table based on the session user_id (class_head_id)
    with connection.cursor() as cursor:
        cursor.execute("SELECT user_id FROM main_classes WHERE id = %s", [class_id])
        logged_in_user = cursor.fetchone()
        if logged_in_user:
            logged_in_user_id = logged_in_user[0]  # Corrected user ID
            print(f"Logged-in User ID: {logged_in_user_id}")


    # Ensure the class head is assigned
    class_head_name = class_details.class_head  
    class_name = class_details.class_name  

    # Fetch subjects linked to the class
    subjects = Subjects.objects.filter(class_obj=class_details)  
    subject_count = subjects.count()  

    # Fetch students for the class
    students = Students.objects.filter(class_obj=class_details)  
    student_count = students.count()

    # Fetch announcements using raw SQL
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, message, created_at FROM main_announcements WHERE class_obj_id = %s ORDER BY created_at DESC", [class_id])
        announcements = cursor.fetchall()

    print("Fetched Announcements:", announcements)  # Debugging output

    # Handle new announcement submission (AJAX request)
    if request.method == "POST" and request.POST.get("announcementText"):
        announcement_text = request.POST.get("announcementText", "").strip()
        if announcement_text:
            with connection.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO main_announcements (message, created_at, class_obj_id) VALUES (%s, NOW(), %s)",
                    [announcement_text, class_id]
                )
                print(f"Inserted Announcement: {announcement_text}")  # Debugging output
            
            return JsonResponse({"success": True, "message": "Announcement added successfully!"})
        else:
            return JsonResponse({"success": False, "message": "Announcement cannot be empty!"})

    return render(request, 'class_head/class_head_class.html', {
        'class_head_name': class_head_name,  
        'class_name': class_name,  
        'students': students,  
        'student_count': student_count,  
        'subjects': subjects,  
        'announcements': announcements,  # Pass announcements to the template
    })




def class_head_profile(request):
    class_head_id = request.session.get('class_id')  # Use the correct session key
    
    if not class_head_id:
        return redirect('class_head_login')  # Redirect if not logged in

    # Fetch the class (which represents the class head) along with its institution
    class_head = get_object_or_404(Classes, id=class_head_id)
    
    # Get institution object using the foreign key relationship (assuming it's a ForeignKey)
    institution = class_head.institution_id  # This is the integer (ID) for institution

    if institution:
        # Fetch the institution object to get the institution_name
        institution_obj = get_object_or_404(Institution, institution_id=institution)
        institution_name = institution_obj.institution_name
    else:
        institution_name = "Unknown Institution"

    if request.method == 'POST':
        class_head.class_head = request.POST.get('name')  # Update class head name
        password = request.POST.get('password')
        if password:  # Only update password if provided
            class_head.password = password
        class_head.save()
        messages.success(request, 'Profile updated successfully.')

    return render(request, 'class_head/class_head_profile.html', {
        'class_head': class_head,
        'institution_name': institution_name
    })



def class_head_chat(request):
    # Retrieve the class_id from session
    class_id = request.session.get('class_id')

    # If class_id is not found in session, redirect to login
    if not class_id:
        return redirect('class_head_login')

    # Retrieve the class head using the class_id
    class_head = Classes.objects.get(id=class_id)
    class_head_user = class_head.user  # Fetch the associated user for the class head

    # Fetch distinct users (subject heads, students, and parents) who can message the logged-in class head
    with connection.cursor() as cursor:
        cursor.execute(""" 
            SELECT DISTINCT u.id, u.role, 
                   CASE
                       WHEN u.role = 'subject_head' THEN s.subject_head
                       WHEN u.role = 'student' THEN st.name
                       WHEN u.role = 'parent' THEN 
                           CASE
                               WHEN p.student_id IS NOT NULL THEN 
                                   (SELECT st.name FROM main_students AS st WHERE st.roll_no = p.student_id)
                               ELSE p.name
                           END
                   END AS user_name
            FROM main_chat AS chat
            JOIN main_users AS u ON (chat.sender_id = u.id OR chat.receiver_id = u.id)
            LEFT JOIN main_subjects AS s ON s.user_id = u.id AND u.role = 'subject_head'
            LEFT JOIN main_students AS st ON st.user_id = u.id AND u.role = 'student'
            LEFT JOIN main_parents AS p ON p.user_id = u.id AND u.role = 'parent'
            WHERE (chat.sender_id = %s OR chat.receiver_id = %s)
            AND u.role != 'class_head'
            ORDER BY user_name;
        """, [class_head_user.id, class_head_user.id])

        involved_users = cursor.fetchall()

        # Debug print to check the data structure
        print("Involved Users:", involved_users)

        # Fetch all students, parents, and subject heads for the class
        cursor.execute("""
            SELECT u.id, u.role, 
                   CASE
                       WHEN u.role = 'subject_head' THEN s.subject_head
                       WHEN u.role = 'student' THEN st.name
                       WHEN u.role = 'parent' THEN 
                           CASE
                               WHEN p.student_id IS NOT NULL THEN 
                                   (SELECT st.name FROM main_students AS st WHERE st.roll_no = p.student_id)
                               ELSE p.name
                           END
                   END AS user_name
            FROM main_users AS u
            LEFT JOIN main_subjects AS s ON s.user_id = u.id AND u.role = 'subject_head'
            LEFT JOIN main_students AS st ON st.user_id = u.id AND u.role = 'student'
            LEFT JOIN main_parents AS p ON p.user_id = u.id AND u.role = 'parent'
            WHERE u.role IN ('subject_head', 'student', 'parent')
            AND (
                st.class_obj_id = %s OR 
                p.student_id IN (SELECT roll_no FROM main_students WHERE class_obj_id = %s) OR
                s.class_obj_id = %s
            )
            ORDER BY user_name;
        """, [class_id, class_id, class_id])

        all_users_for_class = cursor.fetchall()

        # Debug print to check the data structure
        print("All Users for Class:", all_users_for_class)

    # Add users to context for rendering
    context = {
        'class_head': class_head,
        'involved_users': involved_users,
        'all_users_for_class': all_users_for_class,  # New section for class heads to send messages
    }

    return render(request, "class_head/class_head_chat.html", context)



def class_head_chat_user(request, user_id):
    # Retrieve the logged-in class head's ID from session
    class_head_id = request.session.get('class_id')
    if not class_head_id:
        return redirect('class_head_login')

    # Fetch the actual user_id from main_users table based on the session user_id (class_head_id)
    with connection.cursor() as cursor:
        cursor.execute("SELECT user_id FROM main_classes WHERE id = %s", [class_head_id])
        logged_in_user = cursor.fetchone()
        if logged_in_user:
            logged_in_user_id = logged_in_user[0]  # Corrected user ID
            print(f"Logged-in User ID: {logged_in_user_id}")

    # Initialize the user name and role
    selected_user_name = None
    selected_user_role = None

    # Fetch the selected user's name and role
    with connection.cursor() as cursor:
        cursor.execute("SELECT name FROM main_students WHERE user_id = %s", [user_id])
        student_name = cursor.fetchone()
        if student_name:
            selected_user_name = student_name[0]
            selected_user_role = 'Student'
        else:
            cursor.execute("SELECT student_id FROM main_parents WHERE user_id = %s", [user_id])
            parent_student_roll_no = cursor.fetchone()
            if parent_student_roll_no:
                cursor.execute("SELECT name FROM main_students WHERE roll_no = %s", [parent_student_roll_no[0]])
                student_name_for_parent = cursor.fetchone()
                if student_name_for_parent:
                    selected_user_name = f"Parent of {student_name_for_parent[0]}"
                    selected_user_role = 'Parent'
            else:
                cursor.execute("SELECT subject_head FROM main_subjects WHERE user_id = %s", [user_id])
                subject_head_name = cursor.fetchone()
                if subject_head_name:
                    selected_user_name = subject_head_name[0]
                    selected_user_role = 'Subject Head'
                else:
                    cursor.execute("SELECT role FROM main_users WHERE id = %s", [user_id])
                    selected_user = cursor.fetchone()
                    selected_user_role = selected_user[0] if selected_user else 'Unknown'
                    selected_user_name = selected_user_role.capitalize()

    print(f"Chatting with {selected_user_name} ({selected_user_role})")

    # Fetch messages between logged-in user and selected user
    with connection.cursor() as cursor:
        cursor.execute(""" 
            SELECT message, sender_id, receiver_id, created_at 
            FROM main_chat 
            WHERE (sender_id = %s AND receiver_id = %s) 
            OR (sender_id = %s AND receiver_id = %s) 
            ORDER BY created_at;
        """, [user_id, logged_in_user_id, logged_in_user_id, user_id])  
        messages = cursor.fetchall()

    print(f"Messages fetched between {logged_in_user_id} and {user_id}")

    # Handle message sending
    if request.method == 'POST':
        message_text = request.POST.get('message')
        if message_text.strip():  # Ensure message is not empty
            with connection.cursor() as cursor:
                cursor.execute(""" 
                    INSERT INTO main_chat (sender_id, receiver_id, message, created_at) 
                    VALUES (%s, %s, %s, NOW());
                """, [logged_in_user_id, user_id, message_text])
            print(f"Message sent: '{message_text}' from {logged_in_user_id} to {user_id}")
            return redirect('class_head_chat_user', user_id=user_id)  # Refresh page after sending

    context = {
        'selected_user_name': selected_user_name,
        'selected_user_role': selected_user_role,
        'messages': messages,
        'logged_in_user_id': logged_in_user_id,  # Pass correct sender ID
    }
    return render(request, 'class_head/class_head_chat_user.html', context)




def class_head_students(request):
    class_id = request.session.get('class_id')

    if not class_id:
        print("❌ Class ID not found in session. Redirecting to login.")
        return redirect('class_head_login')  # Redirect to login if not logged in

    with connection.cursor() as cursor:
        # Fetch class name
        cursor.execute("""
            SELECT class_name FROM main_classes WHERE id = %s
        """, [class_id])
        class_name = cursor.fetchone()

        if class_name:
            class_name = class_name[0]  # Extract the class name value
            print(f"🏫 Class Name: {class_name}")
        else:
            print(f"⚠️ No class found for class_id: {class_id}")
            class_name = "N/A"

        # Fetch all students for the logged-in class
        cursor.execute("""
            SELECT id, name, roll_no, email FROM main_students
            WHERE class_obj_id = %s
        """, [class_id])
        
        students = cursor.fetchall()

        if not students:
            print(f"⚠️ No students found for class_id: {class_id}")
        else:
            print(f"✅ Found {len(students)} students for class_id: {class_id}")

        # Convert student data to a list of dictionaries
        student_list = []
        for student in students:
            student_data = {
                "id": student[0],
                "name": student[1],
                "roll_no": student[2],
                "email": student[3] if student[3] else "N/A"
            }
            student_list.append(student_data)

        print("📌 Student List:", student_list)  # Debugging: Print the final list

    return render(request, 'class_head/class_head_students.html', {
        "class_name": class_name,
        "students": student_list
    })




def class_head_student_performance(request, student_id):
    # Retrieve the class_id from session
    class_id = request.session.get('class_id')

    if not class_id:
        print("[ERROR] No class_id found in session. Redirecting to class_head login.")
        return redirect('class_head_login')

    subject_id = request.GET.get('subject_id')  # Get the selected subject from query params
    print(f"\n[INFO] Class Head Student Performance View Loaded for Student ID: {student_id}")
    if subject_id:
        print(f"[INFO] Selected Subject ID: {subject_id}")

    student_data = {}
    evaluations = {}
    quiz_percentage = 0  # Default value to avoid None issues
    subjects = []
    selected_subject_name = None

    try:
        with connection.cursor() as cursor:
            # Fetch student details
            cursor.execute("""
                SELECT s.id, s.name, s.roll_no, c.class_name, s.email, c.id AS class_id
                FROM main_students s
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s AND s.class_obj_id = %s
            """, [student_id, class_id])
            row = cursor.fetchone()

            if row:
                student_data = {
                    "id": row[0],  # ✅ Corrected: Student ID
                    "name": row[1],  # ✅ Student Name
                    "roll_no": row[2],
                    "class_name": row[3],
                    "email": row[4] if row[4] else "--",
                    "class_id": row[5]
                }
                print(f"[DEBUG] Student Details: {student_data}")
            else:
                print("[ERROR] Student not found or does not belong to this class.")
                return redirect('class_head_dashboard')

        # Fetch subjects for the student's class
        if student_data.get("class_id"):
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, subject_name FROM main_subjects WHERE class_obj_id = %s
                """, [student_data["class_id"]])
                subjects = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]

            # Get the name of the selected subject
            selected_subject_name = next((s["name"] for s in subjects if str(s["id"]) == subject_id), None)

            if selected_subject_name:
                print(f"[DEBUG] Selected Subject: {selected_subject_name} (ID: {subject_id})")
            else:
                print("[WARNING] Selected subject not found in student's class subjects.")

        # Fetch student evaluation details for the selected subject
        if subject_id:
            print(f"[DEBUG] Fetching evaluation data for subject: {selected_subject_name} (ID: {subject_id})...")
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT marks_percentage, attendance_percentage,
                           study_time_rating, sleep_time_rating, 
                           class_participation_rating, academic_activity_rating
                    FROM main_studentevaluation
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    evaluations = {
                        "marks_percentage": round(row[0] or 0, 2),
                        "attendance_percentage": round(row[1] or 0, 2),
                        "study_time_rating": round(row[2] or 0, 2),
                        "sleep_time_rating": round(row[3] or 0, 2),
                        "class_participation_rating": round(row[4] or 0, 2),
                        "academic_activity_rating": round(row[5] or 0, 2),
                    }
                    print(f"[DEBUG] Evaluation Data: {evaluations}")
                else:
                    print(f"[WARNING] No evaluation data found for {selected_subject_name}.")
        
        # Fetch quiz performance for the selected subject
        if subject_id:
            print(f"[DEBUG] Fetching quiz performance for {selected_subject_name} (ID: {subject_id})...")
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT COUNT(*) AS total_attempted,
                           SUM(CASE WHEN q.correct_option = r.student_response THEN 1 ELSE 0 END) AS correct_answers
                    FROM main_quizresponse r
                    JOIN main_quizquestions q ON r.question_id = q.id
                    JOIN main_quizzes mq ON q.quiz_id = mq.id
                    WHERE r.student_id = %s AND mq.subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    total_attempted, correct_answers = row[0], row[1] or 0
                    quiz_percentage = round((correct_answers / total_attempted) * 100, 2) if total_attempted > 0 else 0
                    print(f"[DEBUG] Quiz Performance: Attempted = {total_attempted}, Correct = {correct_answers}, Percentage = {quiz_percentage}%")
                else:
                    print(f"[WARNING] No quiz data found for {selected_subject_name}.")

        # Prepare data for Graph
        graph_data = [
            float(evaluations.get("marks_percentage", 0)),
            float(evaluations.get("attendance_percentage", 0)),
            float(evaluations.get("study_time_rating", 0)),
            float(evaluations.get("sleep_time_rating", 0)),
            float(evaluations.get("class_participation_rating", 0)),
            float(evaluations.get("academic_activity_rating", 0)),
            float(quiz_percentage)
        ]

    except Exception as e:
        print(f"[ERROR] An error occurred: {e}")
        return redirect('error_page')  # Redirect to an error page if needed

    context = {
        "student": student_data,  # ✅ Now contains `id`
        "evaluations": evaluations,
        "quiz_percentage": quiz_percentage,
        "subjects": subjects,
        "selected_subject_id": int(subject_id) if subject_id else None,
        "graph_data": json.dumps(graph_data)  # ✅ Ensure JSON-safe format
    }

    print("[INFO] Final Context Data Prepared for Rendering.")
    return render(request, "class_head/class_head_student_performance.html", context)





######################################################################################################################


def subject_head_login(request):
    form = SubjectHeadLoginForm()

    if request.method == 'POST':
        form = SubjectHeadLoginForm(request.POST)

        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            # Fetch subject head record
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT s.id, s.password, u.role
                    FROM main_subjects s
                    JOIN main_users u ON s.user_id = u.id
                    WHERE s.email = %s
                """, [email])

                subject_head = cursor.fetchone()

            if subject_head:
                subject_id, db_password, role = subject_head

                # Validate role and password
                if role == 'subject_head' and db_password == password:
                    request.session['subject_id'] = subject_id  # Store subject_id in session
                    return redirect('subject_head_dashboard')  # Redirect to subject head dashboard
                else:
                    messages.error(request, 'Invalid credentials or access restricted.')
            else:
                messages.error(request, 'No Subject Head found with this email.')

    return render(request, 'subject_head/subject_head_login.html', {'form': form})


def subject_head_dashboard(request):
    # Ensure subject head is logged in
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')  # Redirect to login page if not logged in

    # Fetch subject details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT s.subject_name, s.subject_head, c.class_name, c.class_head
            FROM main_subjects s
            JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [subject_id])

        subject_details = cursor.fetchone()  # Fetch one record

    # Prepare context for the template
    context = {
        'subject_name': subject_details[0] if subject_details else None,
        'subject_head': subject_details[1] if subject_details else None,
        'class_name': subject_details[2] if subject_details else None,
        'class_head': subject_details[3] if subject_details else None,
    }

    return render(request, 'subject_head/subject_head_dashboard.html', context)


def subject_head_class(request):
    # Ensure subject head is logged in
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')

    # Fetch class and faculty details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT c.class_name, c.class_head
            FROM main_subjects s
            JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [subject_id])
        class_info = cursor.fetchone()

    if class_info:
        class_name, class_head_name = class_info
    else:
        class_name, class_head_name = "Unknown", "Unknown"

    # Fetch the logged-in subject's details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT subject_name, subject_head, email
            FROM main_subjects
            WHERE id = %s
        """, [subject_id])
        logged_in_subject = cursor.fetchone()

    if logged_in_subject:
        logged_in_subject_name, logged_in_subject_head, logged_in_subject_email = logged_in_subject
    else:
        logged_in_subject_name, logged_in_subject_head, logged_in_subject_email = "Unknown", "Unknown", "Unknown"

    # Fetch subjects under this class
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, subject_name, subject_head, email
            FROM main_subjects
            WHERE class_obj_id = (SELECT class_obj_id FROM main_subjects WHERE id = %s)
        """, [subject_id])
        subjects = [
            {
                'id': row[0],
                'subject_name': row[1],
                'subject_head': row[2],
                'email': row[3],
                'is_logged_in_subject': row[2] == logged_in_subject_head  # Check if the subject belongs to the logged-in user
            }
            for row in cursor.fetchall()
        ]

    # Fetch students under this class
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, roll_no, name, email
            FROM main_students
            WHERE class_obj_id = (SELECT class_obj_id FROM main_subjects WHERE id = %s)
        """, [subject_id])
        students = [
            {'id': row[0], 'roll_no': row[1], 'name': row[2], 'email': row[3]}
            for row in cursor.fetchall()
        ]

    # Get student count
    student_count = len(students)

    context = {
        'class_name': class_name,
        'class_head_name': class_head_name,
        'subjects': subjects,
        'students': students,
        'student_count': student_count,
        'logged_in_subject_head': logged_in_subject_head  # Pass the logged-in subject head to the template
    }

    return render(request, 'subject_head/subject_head_class.html', context)



def subject_head_subjects(request):
    # Ensure subject head is logged in
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')

    # Fetch class and faculty details (for the subject head)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT c.class_name, s.subject_head
            FROM main_subjects s
            JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [subject_id])
        class_info = cursor.fetchone()

    if class_info:
        class_name, subject_head = class_info
    else:
        class_name, subject_head = "Unknown", "Unknown"

    # Fetch assigned subject details for the specific subject head
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, subject_name, subject_head, email
            FROM main_subjects
            WHERE id = %s
        """, [subject_id])
        subjects = [
            {'id': row[0], 'subject_name': row[1], 'subject_head': row[2], 'email': row[3]}
            for row in cursor.fetchall()
        ]

    # Fetch students under this subject's class
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, roll_no, name, email
            FROM main_students
            WHERE class_obj_id = (SELECT class_obj_id FROM main_subjects WHERE id = %s)
        """, [subject_id])
        students = [
            {'id': row[0], 'roll_no': row[1], 'name': row[2], 'email': row[3]}
            for row in cursor.fetchall()
        ]

    # Get student count
    student_count = len(students)

    context = {
        'class_name': class_name,
        'subject_head': subject_head,  # Pass subject head information to the template
        'subjects': subjects,
        'students': students,
        'student_count': student_count
    }

    return render(request, 'subject_head/subject_head_subjects.html', context)


def subject_head_profile(request):
    # Ensure subject head is logged in
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login') 

    if request.method == 'POST':
        # Get updated data from the form
        updated_name = request.POST.get('name')
        updated_password = request.POST.get('password')

        # Update the subject head's details in the database
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE main_subjects
                SET subject_head = %s, password = %s
                WHERE id = %s
            """, [updated_name, updated_password, subject_id])
        
        messages.success(request, "Profile updated successfully!")

        return redirect('subject_head_profile')  # Refresh the page after update

    # Fetch current subject head details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT s.subject_name, s.subject_head, s.email, s.password, c.class_name, c.class_head, i.institution_name
            FROM main_subjects s
            JOIN main_classes c ON s.class_obj_id = c.id
            JOIN main_institution i ON c.institution_id = i.institution_id
            WHERE s.id = %s
        """, [subject_id])

        subject_head = cursor.fetchone()  # Fetch one record

    if subject_head:
        context = {
            'class_head': {
                'subject_name': subject_head[0],  # Subject Name
                'class_head': subject_head[1],  # Subject Head Name
                'email': subject_head[2],
                'password': subject_head[3],
                'class_name': subject_head[4],
            },
            'institution_name': subject_head[6]
        }
    else:
        context = {'class_head': None, 'institution_name': None}

    return render(request, 'subject_head/subject_head_profile.html', context)



def subject_head_chat(request):
    # Ensure subject head is logged in
    subject_id = request.session.get('subject_id')
    if not subject_id:
        print("DEBUG: No subject_id found in session. Redirecting to login.")
        return redirect('subject_head_login')

    print(f"DEBUG: Logged in subject_id: {subject_id}")

    # Fetch user_id and class_id for the logged-in subject head
    with connection.cursor() as cursor:
        cursor.execute("SELECT user_id, class_obj_id FROM main_subjects WHERE id = %s", [subject_id])
        subject_data = cursor.fetchone()

    if not subject_data:
        print(f"DEBUG: No user data found for subject_id {subject_id}.")
        return redirect('subject_head_login')

    subject_user_id, class_id = subject_data
    print(f"DEBUG: Retrieved subject_user_id: {subject_user_id}, class_obj_id: {class_id}")

    # Fetch users involved in chat
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT u.id, u.role, 
            COALESCE(
                c.class_head, 
                s.subject_head, 
                st.name, 
                CASE 
                    WHEN p.user_id IS NOT NULL THEN CONCAT('Parent of ', st2.name) 
                    ELSE NULL 
                END
            ) AS name
        FROM main_users u
        LEFT JOIN main_classes c ON u.id = c.user_id AND u.role = 'class_head'
        LEFT JOIN main_subjects s ON u.id = s.user_id AND u.role = 'subject_head'
        LEFT JOIN main_students st ON u.id = st.user_id AND u.role = 'student'
        LEFT JOIN main_parents p ON u.id = p.user_id AND u.role = 'parent'
        LEFT JOIN main_students st2 ON p.student_id = st2.roll_no AND u.role = 'parent'
        WHERE u.id IN (
            SELECT DISTINCT sender_id FROM main_chat WHERE receiver_id = %s
            UNION
            SELECT DISTINCT receiver_id FROM main_chat WHERE sender_id = %s
        );
        """, [subject_user_id, subject_user_id])

        chat_users = [{'id': row[0], 'role': row[1], 'name': row[2] or 'Unknown'} for row in cursor.fetchall()]

    print(f"DEBUG: Retrieved users involved in chat: {chat_users}")

    # Fetch users the subject head can send messages to (students, parents, class head, other subject heads)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT u.id, u.role, 
            COALESCE(
                st.name, 
                CASE 
                    WHEN p.user_id IS NOT NULL THEN CONCAT('Parent of ', st2.name) 
                    ELSE NULL 
                END,
                c.class_head,
                s.subject_head
            ) AS name
        FROM main_users u
        LEFT JOIN main_students st ON u.id = st.user_id AND u.role = 'student' AND st.class_obj_id = %s
        LEFT JOIN main_parents p ON u.id = p.user_id AND u.role = 'parent'  
        LEFT JOIN main_students st2 ON p.student_id = st2.roll_no AND u.role = 'parent' AND st2.class_obj_id = %s
        LEFT JOIN main_classes c ON u.id = c.user_id AND u.role = 'class_head' AND c.id = %s
        LEFT JOIN main_subjects s ON u.id = s.user_id AND u.role = 'subject_head' AND s.class_obj_id = %s
        WHERE (st.class_obj_id IS NOT NULL OR st2.class_obj_id IS NOT NULL OR c.id IS NOT NULL OR (s.class_obj_id IS NOT NULL AND u.id != %s));
        """, [class_id, class_id, class_id, class_id, subject_user_id])

        message_users = [{'id': row[0], 'role': row[1], 'name': row[2] or 'Unknown'} for row in cursor.fetchall()]

    print(f"DEBUG: Retrieved users the subject head can send messages to: {message_users}")

    return render(request, 'subject_head/subject_head_chat.html', {'chat_users': chat_users, 'message_users': message_users})



def subject_head_chat_user(request, user_id):
    # Ensure subject head is logged in
    subject_id = request.session.get('subject_id')
    if not subject_id:
        print("DEBUG: No subject_id found in session. Redirecting to login.")
        return redirect('subject_head_login')

    print(f"DEBUG: Logged in subject_id: {subject_id}")

    # Fetch user_id for the logged-in subject head
    with connection.cursor() as cursor:
        cursor.execute("SELECT user_id FROM main_subjects WHERE id = %s", [subject_id])
        subject_user_id = cursor.fetchone()

    if not subject_user_id:
        print(f"DEBUG: No user_id found for subject_id {subject_id}. Redirecting to login.")
        return redirect('subject_head_login')

    subject_user_id = subject_user_id[0]
    print(f"DEBUG: Retrieved subject_user_id: {subject_user_id}")

    # Fetch selected user's name and role
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT u.role, 
            COALESCE(
                c.class_head, 
                s.subject_head, 
                st.name, 
                CASE 
                    WHEN p.user_id IS NOT NULL THEN CONCAT('Parent of ', st2.name) 
                    ELSE NULL 
                END
            ) AS name
        FROM main_users u
        LEFT JOIN main_classes c ON u.id = c.user_id AND u.role = 'class_head'
        LEFT JOIN main_subjects s ON u.id = s.user_id AND u.role = 'subject_head'
        LEFT JOIN main_students st ON u.id = st.user_id AND u.role = 'student'
        LEFT JOIN main_parents p ON u.id = p.user_id AND u.role = 'parent'
        LEFT JOIN main_students st2 ON p.student_id = st2.roll_no AND u.role = 'parent'
        WHERE u.id = %s
        """, [user_id])

        selected_user = cursor.fetchone()

    if not selected_user:
        print(f"DEBUG: No user found with id {user_id}. Redirecting to chat page.")
        return redirect('subject_head_chat')

    selected_user_role, selected_user_name = selected_user
    print(f"DEBUG: Retrieved selected_user_role: {selected_user_role}, selected_user_name: {selected_user_name}")

    # Handle message sending
    if request.method == "POST":
        message_text = request.POST.get("message", "").strip()
        
        if message_text:
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO main_chat (message, sender_id, receiver_id, created_at) 
                    VALUES (%s, %s, %s, %s)
                """, [message_text, subject_user_id, user_id, now()])
            
            print(f"DEBUG: Message sent from {subject_user_id} to {user_id}: {message_text}")
            return redirect('subject_head_chat_user', user_id=user_id)
        else:
            print("DEBUG: Empty message, not sent.")

    # Fetch chat messages between subject head and selected user
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT message, sender_id, receiver_id, created_at 
            FROM main_chat 
            WHERE (sender_id = %s AND receiver_id = %s) 
               OR (sender_id = %s AND receiver_id = %s) 
            ORDER BY created_at ASC
        """, [subject_user_id, user_id, user_id, subject_user_id])

        messages = cursor.fetchall()

    print(f"DEBUG: Retrieved {len(messages)} messages between subject_head_user_id {subject_user_id} and user_id {user_id}")

    return render(request, 'subject_head/subject_head_chat_user.html', {
        'selected_user_role': selected_user_role,
        'selected_user_name': selected_user_name,
        'messages': messages,
        'logged_in_user_id': subject_user_id
    })


@csrf_exempt
def subject_head_quiz(request):
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')

    # Fetch subject details
    with connection.cursor() as cursor:
        cursor.execute("SELECT user_id, subject_name FROM main_subjects WHERE id = %s", [subject_id])
        logged_in_user = cursor.fetchone()
        logged_in_user_id = logged_in_user[0] if logged_in_user else None
        subject_name = logged_in_user[1] if logged_in_user else "Unknown Subject"

    # Fetch assigned classes
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT c.id, c.class_name 
            FROM main_classes c
            JOIN main_subjects s ON c.id = s.class_obj_id
            WHERE s.id = %s
        """, [subject_id])
        classes = cursor.fetchall()

    # Fetch quizzes created by this subject head
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT q.id, q.name, c.class_name 
            FROM main_quizzes q
            JOIN main_classes c ON q.class_obj_id = c.id
            WHERE q.subject_id = %s
        """, [subject_id])
        quizzes = cursor.fetchall()

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')

            if action == "create_quiz":
                quiz_name = data.get('quiz_name')
                class_id = data.get('class_id')

                if not (quiz_name and class_id):
                    return JsonResponse({'error': 'Missing required fields'}, status=400)

                with connection.cursor() as cursor:
                    cursor.execute(
                        "INSERT INTO main_quizzes (name, subject_id, class_obj_id) VALUES (%s, %s, %s)",
                        [quiz_name, subject_id, class_id]
                    )
                    cursor.execute("SELECT LAST_INSERT_ID()")
                    quiz_id = cursor.fetchone()[0]

                return JsonResponse({'quiz_id': quiz_id, 'message': 'Quiz created successfully'})

            elif action == "add_questions":
                questions = data.get('questions')

                if not questions or not isinstance(questions, list):
                    return JsonResponse({'error': 'Invalid questions data'}, status=400)

                with connection.cursor() as cursor:
                    for question in questions:
                        cursor.execute("""
                            INSERT INTO main_quizquestions (quiz_id, question, option_a, option_b, option_c, option_d, correct_option) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, [
                            question['quiz_id'],
                            question['question'],
                            question['option_a'],
                            question['option_b'],
                            question['option_c'],
                            question['option_d'],
                            question['correct_option']
                        ])

                return JsonResponse({'message': 'Questions added successfully'})

            else:
                return JsonResponse({'error': 'Invalid action'}, status=400)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            print(f"ERROR: {e}")
            return JsonResponse({'error': str(e)}, status=500)
        
    # Fetch quizzes with the number of questions
    quizzes = []
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT mq.id, mq.name, c.class_name, COUNT(mqq.id) 
            FROM main_quizzes mq
            JOIN main_classes c ON mq.class_obj_id = c.id
            LEFT JOIN main_quizquestions mqq ON mq.id = mqq.quiz_id
            WHERE mq.subject_id = %s
            GROUP BY mq.id, mq.name, c.class_name
        """, [subject_id])
        quizzes = cursor.fetchall()

    return render(request, 'subject_head/subject_head_quiz.html', {
        'classes': classes,
        'quizzes': quizzes,
        'subject_name': subject_name,
        'subject_id': subject_id,
        "quizzes": quizzes,
    })


def subject_head_studys(request):
    # Ensure the user is a subject head
    subject_id = request.session.get('subject_id')
    if not subject_id:
        print("Redirecting: No subject_id found in session.")
        return redirect('subject_head_login')

    print(f"Session subject_id: {subject_id}")

    # Fetch class_obj_id for the logged-in subject head
    with connection.cursor() as cursor:
        cursor.execute("SELECT class_obj_id FROM main_subjects WHERE id = %s", [subject_id])
        class_obj = cursor.fetchone()
    
    class_obj_id = class_obj[0] if class_obj else None
    print(f"Fetched class_obj_id: {class_obj_id}")

    # Fetch uploaded documents for this subject and class
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, file_url, announcement, created_at 
            FROM main_studymaterials 
            WHERE subject_id = %s AND class_obj_id = %s
            ORDER BY created_at DESC
        """, [subject_id, class_obj_id])
        docs = cursor.fetchall()
    
    print(f"Fetched {len(docs)} documents for subject_id: {subject_id}, class_obj_id: {class_obj_id}")

    if request.method == "POST":
        print("POST request received.")

        announcement = request.POST.get('announcement', '')
        file = request.FILES.get('file_url')  # Corrected file field name

        print(f"Announcement: {announcement}")
        print(f"Received FILES: {request.FILES}")  # Debugging file upload
        print(f"File received: {file.name if file else 'No file uploaded'}")

        file_url = None  # Default to None

        # Handle file upload
        if file and class_obj_id and subject_id:
            upload_dir = "media/uploads/"
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir)
                print(f"Created directory: {upload_dir}")

            file_path = os.path.join(upload_dir, file.name)

            with open(file_path, 'wb+') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)

            print(f"File saved successfully at: {file_path}")
            file_url = file_path  # Store the file path in the database

        # Insert data into the database
        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO main_studymaterials (file_url, created_at, class_obj_id, subject_id, announcement) 
                VALUES (%s, %s, %s, %s, %s)
            """, [file_url, datetime.datetime.now(), class_obj_id, subject_id, announcement if announcement else None])

        print("Database entry created for study material.")
        messages.success(request, "Study material uploaded successfully!")
        return redirect('subject_head_studys')

    print("Rendering template with uploaded documents.")
    return render(request, 'subject_head/subject_head_study.html', {'docs': docs})




def subject_head_marks(request):
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')

    subject_data = None
    student_list = []

    try:
        with connection.cursor() as cursor:
            # Fetch subject details along with class name
            cursor.execute("""
                SELECT s.id, s.subject_name, c.class_name, s.class_obj_id 
                FROM main_subjects s
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s
            """, [subject_id])
            subject = cursor.fetchone()

            if subject:
                subject_data = {
                    'id': subject[0],
                    'subject_name': subject[1],
                    'class_obj_name': subject[2],  # Class name
                    'class_obj_id': subject[3],
                }

                # Handle form submission to update marks
                if request.method == "POST":
                    success = True  # Track success

                    for key, value in request.POST.items():
                        if key.startswith("marks_"):
                            student_id = key.split("_")[1]
                            mark_value = value.strip()

                            try:
                                mark_value = float(mark_value)  # Convert to float safely

                                # ✅ Step 1: Check if the record exists in `main_marks`
                                cursor.execute("""
                                    SELECT id FROM main_marks WHERE student_id = %s AND subject_id = %s
                                """, [student_id, subject_data['id']])
                                mark_record = cursor.fetchone()

                                if mark_record:
                                    # ✅ Step 2: If exists, UPDATE `main_marks`
                                    cursor.execute("""
                                        UPDATE main_marks 
                                        SET mark_percentage = %s 
                                        WHERE student_id = %s AND subject_id = %s
                                    """, [mark_value, student_id, subject_data['id']])
                                else:
                                    # ✅ Step 3: If not exists, INSERT into `main_marks`
                                    cursor.execute("""
                                        INSERT INTO main_marks (mark_percentage, student_id, subject_id)
                                        VALUES (%s, %s, %s)
                                    """, [mark_value, student_id, subject_data['id']])

                                # ✅ Step 4: Check if the record exists in `main_studentevaluation`
                                cursor.execute("""
                                    SELECT id FROM main_studentevaluation
                                    WHERE student_id = %s AND subject_id = %s
                                """, [student_id, subject_data['id']])
                                evaluation = cursor.fetchone()

                                if evaluation:
                                    # ✅ Step 5: If exists, UPDATE `main_studentevaluation`
                                    cursor.execute("""
                                        UPDATE main_studentevaluation 
                                        SET marks_percentage = %s
                                        WHERE student_id = %s AND subject_id = %s
                                    """, [mark_value, student_id, subject_data['id']])
                                else:
                                    # ✅ Step 6: If not exists, INSERT into `main_studentevaluation`
                                    cursor.execute("""
                                        INSERT INTO main_studentevaluation (student_id, subject_id, marks_percentage)
                                        VALUES (%s, %s, %s)
                                    """, [student_id, subject_data['id'], mark_value])

                            except ValueError:  # Catch invalid input
                                print(f"Invalid mark value for student {student_id}: {value}")
                                success = False

                    # Display success or error message
                    if success:
                        messages.success(request, "✅ Marks updated successfully!", extra_tags='marks_success')
                    else:
                        messages.error(request, "❌ Failed to update marks. Please try again.", extra_tags='marks_error')

                # Fetch students of this class
                cursor.execute("SELECT id, roll_no, name FROM main_students WHERE class_obj_id = %s", [subject_data['class_obj_id']])
                students = cursor.fetchall()

                for student in students:
                    student_id = student[0]

                    # Fetch marks for each student
                    cursor.execute("""
                        SELECT mark_percentage FROM main_marks 
                        WHERE student_id = %s AND subject_id = %s
                    """, [student_id, subject_data['id']])
                    mark = cursor.fetchone()

                    student_list.append({
                        'id': student_id,
                        'roll_no': student[1],
                        'name': student[2],
                        'mark_percentage': mark[0] if mark else None
                    })

    except Exception as e:
        print(f"Database error: {e}")
        messages.error(request, "An unexpected error occurred.", extra_tags='marks_error')

    return render(request, 'subject_head/subject_head_marks.html', {
        'subject': subject_data,
        'students': student_list
    })



def download_marks_template(request, subject_id):
    print(f"🔍 Fetching details for Subject ID: {subject_id}")

    with connection.cursor() as cursor:
        # Fetch subject name and class ID
        cursor.execute("""
            SELECT subject_name, class_obj_id FROM main_subjects WHERE id = %s
        """, [subject_id])
        subject_data = cursor.fetchone()

        if not subject_data:
            print("❌ Invalid subject ID! No data found.")
            return HttpResponse("Invalid subject ID", status=400)

        subject_name, class_id = subject_data
        print(f"✅ Retrieved Subject Name: {subject_name}, Class ID: {class_id}")

        # Fetch students belonging to the same class as the subject
        cursor.execute("""
            SELECT s.roll_no, s.name
            FROM main_students s
            WHERE s.class_obj_id = %s
        """, [class_id])

        students_data = cursor.fetchall()
        print(f"✅ Found {len(students_data)} students in Class ID: {class_id}")

    # Create a new Excel workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks Template"
    print("✅ Created new Excel workbook.")

    # Define column headers
    headers = ["Roll No", "Student Name", "Subject Name", "Mark Percentage"]
    ws.append(headers)
    print("✅ Added headers:", headers)

    # Store Excel data for debugging
    excel_data = [headers]

    # Write student data (mark_percentage column is left empty)
    for roll_no, name in students_data:
        row = [roll_no, name, subject_name, ""]
        ws.append(row)
        excel_data.append(row)

    print("✅ Populated Excel sheet with student details.")

    # Print final contents of the dynamically created Excel
    print("\n📋 Final Excel Content:")
    for row in excel_data:
        print(row)

    # Prepare the response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Marks_Template_{subject_name}.xlsx"'

    # Save the workbook to response
    wb.save(response)
    print("✅ Excel file generated and sent as response.")

    return response



def upload_marks(request):
    if request.method == "POST" and request.FILES.get("file"):
        excel_file = request.FILES["file"]
        print(f"📂 Received file: {excel_file.name}")

        try:
            # Open the uploaded Excel file
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            print(f"✅ Opened Excel file. Sheet: {ws.title}")

            # Read headers
            headers = [str(cell.value).strip() for cell in ws[1] if cell.value]
            print(f"📑 Extracted Headers: {headers}")

            expected_headers = ["Roll No", "Student Name", "Subject Name", "Mark Percentage"]

            # Case-insensitive header validation
            if [h.lower() for h in headers] != [eh.lower() for eh in expected_headers]:
                print(f"❌ Invalid file format. Headers do not match. Expected: {expected_headers}, Found: {headers}")
                messages.error(request, "❌ Invalid file format! Please use the correct template.", extra_tags="marks_error")
                return redirect("subject_head_marks")

            print("✅ Headers validated. Processing rows...")

            # Fetch subject_id from session
            subject_id = request.session.get("subject_id")
            if not subject_id:
                print("❌ Subject ID not found in session. Redirecting to login.")
                messages.error(request, "❌ Please log in again.", extra_tags="marks_error")
                return redirect("subject_head_login")

            print(f"🎯 Subject ID identified from session: {subject_id}")

            rows_processed = 0

            # Start transaction for atomic inserts/updates
            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_values = [cell for cell in row if cell is not None]  # Remove None values

                    # Ensure we have exactly 4 values (skip rows with extra/missing columns)
                    if len(row_values) != 4:
                        print(f"⚠️ Skipping row due to incorrect column count: {row_values}")
                        continue

                    roll_no, student_name, subject_name, mark_percentage = row_values
                    print(f"🔍 Processing: Roll No: {roll_no}, Student: {student_name}, Subject: {subject_name}, Marks: {mark_percentage}")

                    # Validate mark_percentage
                    if mark_percentage is None or not isinstance(mark_percentage, (int, float)):
                        print(f"⚠️ Skipping Roll No {roll_no} - Invalid marks data.")
                        continue

                    with connection.cursor() as cursor:
                        # Get student_id using roll_no
                        cursor.execute("SELECT id FROM main_students WHERE roll_no = %s", [roll_no])
                        student_data = cursor.fetchone()

                        if not student_data:
                            print(f"❌ Skipping Roll No {roll_no} - Student not found.")
                            continue

                        student_id = student_data[0]

                        # Try updating existing marks in main_marks
                        cursor.execute("""
                            UPDATE main_marks 
                            SET mark_percentage = %s 
                            WHERE student_id = %s AND subject_id = %s
                        """, [mark_percentage, student_id, subject_id])

                        if cursor.rowcount > 0:
                            print(f"🔄 Updated marks for Roll No {roll_no}: {mark_percentage}%")
                        else:
                            # If no update, insert new marks
                            cursor.execute("""
                                INSERT INTO main_marks (student_id, subject_id, mark_percentage) 
                                VALUES (%s, %s, %s)
                            """, [student_id, subject_id, mark_percentage])
                            print(f"🆕 Inserted marks for Roll No {roll_no}: {mark_percentage}%")

                        # Now update main_studentevaluation for the same student and subject
                        cursor.execute("""
                            UPDATE main_studentevaluation 
                            SET marks_percentage = %s 
                            WHERE student_id = %s AND subject_id = %s
                        """, [mark_percentage, student_id, subject_id])

                        if cursor.rowcount > 0:
                            print(f"🔄 Updated StudentEvaluation for Roll No {roll_no}: {mark_percentage}%")
                        else:
                            # If no update, insert new record
                            cursor.execute("""
                                INSERT INTO main_studentevaluation (student_id, subject_id, marks_percentage) 
                                VALUES (%s, %s, %s)
                            """, [student_id, subject_id, mark_percentage])
                            print(f"🆕 Inserted StudentEvaluation for Roll No {roll_no}: {mark_percentage}%")

                    rows_processed += 1

            print(f"✅ Finished processing {rows_processed} rows.")
            messages.success(request, f"✅ Successfully updated {rows_processed} student marks.", extra_tags="marks_success")

        except Exception as e:
            print(f"❌ Error processing file: {e}")
            messages.error(request, "❌ An error occurred while processing the file. Please try again.", extra_tags="marks_error")

        return redirect("subject_head_marks")

    print("❌ No file uploaded or invalid request method.")
    messages.error(request, "Please upload a valid Excel file.")
    return redirect("subject_head_marks")



def subject_head_attendance(request):
    subject_id = request.session.get('subject_id')

    if not subject_id:
        print("❌ Subject ID not found in session. Redirecting to login.")
        return redirect('subject_head_login')

    with connection.cursor() as cursor:
        # Fetch subject and class details
        cursor.execute("""
            SELECT s.subject_name, c.class_name 
            FROM main_subjects s
            JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [subject_id])
        subject = cursor.fetchone()

        if not subject:
            print(f"❌ No subject found for subject_id: {subject_id}. Redirecting to dashboard.")
            return redirect('subject_head_dashboard')

        subject_name, class_name = subject  
        print(f"✅ Subject: {subject_name}, Class: {class_name}")

        # Fetch students of the subject's class
        cursor.execute("""
            SELECT s.id, s.roll_no, s.name 
            FROM main_students s
            WHERE s.class_obj_id = (SELECT class_obj_id FROM main_subjects WHERE id = %s)
        """, [subject_id])
        students = cursor.fetchall()
        print(f"📌 Students Fetched ({len(students)} students): {students}")

        # Handle POST request (Attendance Submission)
        if request.method == "POST":
            print("📝 Received Attendance Submission")

            attendance_date = request.POST.get('attendance_date')
            hour = request.POST.get('hour')

            if not attendance_date or not hour:
                print("⚠️ Missing Date or Hour. Showing error message.")
                messages.error(request, "⚠️ Please select both date and hour.")
                return redirect('subject_head_attendance')

            hour = int(hour)  # Convert hour to integer

            # **Check if attendance already exists for this date & hour**
            cursor.execute("""
                SELECT COUNT(*) FROM main_attendance 
                WHERE subject_id = %s AND attendance_date = %s AND hour = %s
            """, [subject_id, attendance_date, hour])
            existing_records = cursor.fetchone()[0]

            if existing_records > 0:
                print(f"🚨 Attendance for {attendance_date}, Hour {hour} already exists!")
                messages.warning(request, f"⚠️ Attendance for {attendance_date}, Hour {hour} has already been recorded.")
                return redirect('subject_head_attendance')

            # Insert new attendance records
            for student in students:
                student_id = student[0]
                status = request.POST.get(f'attendance_{student_id}', 'absent')  # Default to 'absent' if not checked

                cursor.execute("""
                    INSERT INTO main_attendance (student_id, subject_id, attendance_date, hour, status, created_at)
                    VALUES (%s, %s, %s, %s, %s, NOW())
                """, [student_id, subject_id, attendance_date, hour, status])
                print(f"✅ Attendance Added - Student: {student_id}, Date: {attendance_date}, Hour: {hour}, Status: {status}")

                # **Calculate Attendance Percentage**
                cursor.execute("""
                    SELECT COUNT(*) FROM main_attendance 
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                total_classes = cursor.fetchone()[0]

                cursor.execute("""
                    SELECT COUNT(*) FROM main_attendance 
                    WHERE student_id = %s AND subject_id = %s AND status = 'present'
                """, [student_id, subject_id])
                present_count = cursor.fetchone()[0]

                attendance_percentage = (present_count / total_classes) * 100 if total_classes > 0 else 0
                print(f"📊 Student {student_id} - Attendance: {attendance_percentage:.2f}%")

                # **Update or Insert into StudentEvaluation**
                cursor.execute("""
                    SELECT id FROM main_studentevaluation 
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                existing_record = cursor.fetchone()

                if existing_record:
                    cursor.execute("""
                        UPDATE main_studentevaluation 
                        SET attendance_percentage = %s 
                        WHERE student_id = %s AND subject_id = %s
                    """, [attendance_percentage, student_id, subject_id])
                    print(f"🔄 Updated Attendance Percentage: {attendance_percentage:.2f}% for Student {student_id}")
                else:
                    cursor.execute("""
                        INSERT INTO main_studentevaluation (student_id, subject_id, attendance_percentage)
                        VALUES (%s, %s, %s)
                    """, [student_id, subject_id, attendance_percentage])
                    print(f"✅ Inserted New Attendance Percentage: {attendance_percentage:.2f}% for Student {student_id}")

            messages.success(request, "✅ Attendance successfully recorded!")
            return redirect('subject_head_attendance')

    context = {
        'subject_name': subject_name,
        'class_name': class_name,
        'students': students,
    }

    return render(request, 'subject_head/subject_head_attendance.html', context)


def subject_head_attendance_history(request):
    subject_id = request.session.get('subject_id')

    if not subject_id:
        print("❌ Subject ID not found in session. Redirecting to login.")
        return redirect('subject_head_login')

    print(f"✅ Subject ID found: {subject_id}")

    # Get filter parameters from request
    filter_date = request.GET.get('date', '').strip()
    filter_student = request.GET.get('student', '').strip()

    print(f"🔍 Filter Date: {filter_date}, Filter Student: {filter_student}")

    # Base SQL query with correct table name, ordered by date
    sql_query = """
        SELECT 
            a.id, 
            a.attendance_date AS date, 
            a.hour, 
            s.roll_no, 
            COALESCE(s.name, '') AS student_name,  -- Handle NULL names
            a.status
        FROM main_attendance a
        JOIN main_students s ON a.student_id = s.id
        JOIN main_subjects sub ON a.subject_id = sub.id
        WHERE sub.id = %s
    """
    
    sql_params = [subject_id]

    # Apply filters dynamically
    if filter_date:
        sql_query += " AND a.attendance_date = %s"
        sql_params.append(filter_date)

    if filter_student:
        sql_query += " AND COALESCE(s.name, '') LIKE %s"  # Avoid NULL issues
        sql_params.append(f"%{filter_student}%")

    # Order by date (newest first) and hour (ascending)
    sql_query += " ORDER BY a.attendance_date DESC, a.hour ASC"

    # Execute raw SQL query
    with connection.cursor() as cursor:
        cursor.execute(sql_query, sql_params)
        rows = cursor.fetchall()

    print(f"✅ Retrieved {len(rows)} attendance records.")

    # Convert query result to a list of dictionaries
    attendance_records = [
        {"id": row[0], "date": row[1], "hour": row[2], "roll_no": row[3], "student_name": row[4], "status": row[5]}
        for row in rows
    ]

    return render(request, 'subject_head/subject_head_attendance_history.html', {
        "attendance_records": attendance_records,
        "filter_date": filter_date,
        "filter_student": filter_student,
    })



def update_attendance(request):
    if request.method == "POST":
        attendance_ids = request.POST.getlist("attendance_ids[]")  # List of attendance record IDs
        statuses = request.POST.getlist("status[]")  # List of new statuses

        print("\n--- DEBUG: Incoming Attendance Update Request ---")
        print(f"Attendance IDs: {attendance_ids}")
        print(f"New Statuses: {statuses}")

        try:
            with connection.cursor() as cursor:
                for att_id, status in zip(attendance_ids, statuses):
                    # Update attendance status in the database
                    cursor.execute("""
                        UPDATE main_attendance
                        SET status = %s
                        WHERE id = %s
                    """, [status, att_id])
                    print(f"✅ Updated attendance ID {att_id} to status: {status}")

                # Step 1: Get all unique student IDs that were updated
                cursor.execute("""
                    SELECT DISTINCT student_id FROM main_attendance WHERE id IN %s
                """, [tuple(attendance_ids)])
                student_ids = [row[0] for row in cursor.fetchall()]

                print(f"\n🔍 Students affected: {student_ids}")

                for student_id in student_ids:
                    # Step 2: Recalculate attendance percentage
                    cursor.execute("""
                        SELECT 
                            SUM(CASE WHEN status = 'present' THEN 1 ELSE 0 END) * 100.0 / COUNT(*)
                        FROM main_attendance
                        WHERE student_id = %s
                    """, [student_id])
                    attendance_percentage = cursor.fetchone()[0] or 0  # Ensure default value if no records found

                    print(f"📊 Student {student_id} new attendance percentage: {attendance_percentage:.2f}%")

                    # Step 3: Update `student_evaluations` table
                    cursor.execute("""
                        UPDATE main_studentevaluation
                        SET attendance_percentage = %s
                        WHERE student_id = %s
                    """, [attendance_percentage, student_id])

                    print(f"✅ Updated student_evaluations for student {student_id}")

            messages.success(request, "✅ Attendance updated successfully!")
        except Exception as e:
            messages.error(request, f"❌ Error updating attendance: {str(e)}")
            print(f"\n❌ ERROR: {str(e)}")

        return redirect("subject_head_attendance_history")

    messages.error(request, "❌ Invalid request.")
    return redirect("update_attendance")



def delete_attendance(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            record_id = data.get("id")

            if not record_id:
                return JsonResponse({"success": False, "error": "Invalid record ID."})

            print(f"Deleting attendance record ID: {record_id}")

            with connection.cursor() as cursor:
                # Fetch student_id and subject_id before deleting
                cursor.execute("SELECT student_id, subject_id FROM main_attendance WHERE id = %s", [record_id])
                result = cursor.fetchone()

                if not result:
                    return JsonResponse({"success": False, "error": "Attendance record not found."})

                student_id, subject_id = result
                print(f"Found record - Student ID: {student_id}, Subject ID: {subject_id}")

                # Delete the attendance record
                cursor.execute("DELETE FROM main_attendance WHERE id = %s", [record_id])
                print(f"Attendance record {record_id} deleted successfully.")

                # Calculate new attendance percentage
                cursor.execute("""
                    SELECT COUNT(*) FROM main_attendance
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                total_classes = cursor.fetchone()[0]

                cursor.execute("""
                    SELECT COUNT(*) FROM main_attendance
                    WHERE student_id = %s AND subject_id = %s AND status = 'present'
                """, [student_id, subject_id])
                total_present = cursor.fetchone()[0]

                if total_classes > 0:
                    new_attendance_percentage = (total_present / total_classes) * 100
                else:
                    new_attendance_percentage = 0  # No attendance records left

                print(f"Updated Attendance Percentage: {new_attendance_percentage}%")

                # Update the student_evaluations table
                cursor.execute("""
                    UPDATE main_studentevaluation
                    SET attendance_percentage = %s
                    WHERE student_id = %s AND subject_id = %s
                """, [new_attendance_percentage, student_id, subject_id])

                print(f"Attendance percentage updated for Student ID {student_id}.")

            return JsonResponse({"success": True})

        except Exception as e:
            print(f"Error deleting attendance: {e}")
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method."})




def subject_head_evaluation(request):
    subject_id = request.session.get('subject_id')

    if not subject_id:
        print("❌ Subject ID not found in session. Redirecting to login.")
        return redirect('subject_head_login')

    # Fetch students for the given subject
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT main_students.id, main_students.name, main_students.roll_no
            FROM main_students
            JOIN main_classes ON main_students.class_obj_id = main_classes.id
            JOIN main_subjects ON main_classes.id = main_subjects.class_obj_id
            WHERE main_subjects.id = %s
        """, [subject_id])
        students = cursor.fetchall()

    student_list = []
    student_ids = [row[0] for row in students]
    ratings_dict = {}

    if student_ids:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT student_id, academic_activity_rating, class_participation_rating
                FROM main_studentevaluation
                WHERE subject_id = %s AND student_id IN %s
            """, [subject_id, tuple(student_ids)])

            ratings = cursor.fetchall()
            for rating in ratings:
                ratings_dict[rating[0]] = {
                    "academic_activity_rating": rating[1] if rating[1] is not None else 0,
                    "class_participation_rating": rating[2] if rating[2] is not None else 0
                }

    for row in students:
        student_id, name, roll_no = row
        student_list.append({
            'id': student_id,
            'name': name,
            'roll_no': roll_no,
            'academic_activity_rating': ratings_dict.get(student_id, {}).get("academic_activity_rating", 0),
            'class_participation_rating': ratings_dict.get(student_id, {}).get("class_participation_rating", 0),
        })

    # Handle Rating Submission
    if request.method == "POST":
        rating_type = request.POST.get("rating_type")  # Should be 'academic_activity_rating' or 'class_participation_rating'

        # ✅ Debugging: Print received data to check if values exist
        print(f"🔹 Received rating type: {rating_type}")
        print(f"🔹 POST Data: {request.POST}")

        # Extract rating inputs from POST data
        student_ratings = {
            key.split("_")[-1]: float(value) for key, value in request.POST.items()
            if key.startswith("class_participation_rating_") or key.startswith("academic_activity_rating_")
        }

        print(f"✅ Parsed student ratings: {student_ratings}")  # Debugging output

        # Ensure all students are updated even if their rating is unchanged
        with connection.cursor() as cursor:
            for student in student_list:
                student_id = student['id']
                roll_no = student['roll_no']

                # Get the updated value or keep the existing one
                new_rating = student_ratings.get(str(roll_no), student[rating_type])

                print(f"🔄 Updating {rating_type} for Student {roll_no} (ID: {student_id}) to {new_rating}")  # Debugging

                # Check if record exists
                cursor.execute("""
                    SELECT id FROM main_studentevaluation
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                record = cursor.fetchone()

                if record:
                    # Update existing record
                    cursor.execute(f"""
                        UPDATE main_studentevaluation
                        SET {rating_type} = %s
                        WHERE student_id = %s AND subject_id = %s
                    """, [new_rating, student_id, subject_id])
                else:
                    # Insert new record
                    cursor.execute(f"""
                        INSERT INTO main_studentevaluation (student_id, subject_id, {rating_type})
                        VALUES (%s, %s, %s)
                    """, [student_id, subject_id, new_rating])

        messages.success(request, f"{rating_type.replace('_', ' ').title()} updated successfully!")

        # ✅ Redirect to prevent form resubmission
        return redirect('subject_head_evaluation')

    return render(request, 'subject_head/subject_head_evaluation.html', {'students': student_list})




def subject_head_students(request, student_id):
    subject_id = request.session.get('subject_id')

    if not subject_id:
        print("❌ Subject ID not found in session. Redirecting to login.")
        return redirect('subject_head_login')

    with connection.cursor() as cursor:
        # Fetch student details
        cursor.execute("""
            SELECT s.name, s.roll_no, c.class_name, s.email 
            FROM main_students s
            LEFT JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [student_id])
        student = cursor.fetchone()

        if not student:
            return redirect('subject_head_dashboard')

        student_data = {
            "name": student[0],
            "roll_no": student[1],
            "class_name": student[2] if student[2] else "N/A",
            "email": student[3] if student[3] else "N/A"
        }

        # Fetch evaluation details
        cursor.execute("""
            SELECT study_time_rating, sleep_time_rating, class_participation_rating, 
                   academic_activity_rating, attendance_percentage, marks_percentage
            FROM main_studentevaluation
            WHERE student_id = %s AND subject_id = %s
        """, [student_id, subject_id])
        evaluation = cursor.fetchone()

        evaluations = {
            "study_time_rating": round(evaluation[0], 2) if evaluation and evaluation[0] is not None else "--",
            "sleep_time_rating": round(evaluation[1], 2) if evaluation and evaluation[1] is not None else "--",
            "class_participation_rating": round(evaluation[2], 2) if evaluation and evaluation[2] is not None else "--",
            "academic_activity_rating": round(evaluation[3], 2) if evaluation and evaluation[3] is not None else "--",
            "attendance_percentage": round(evaluation[4], 2) if evaluation and evaluation[4] is not None else "--",
            "marks_percentage": round(evaluation[5], 2) if evaluation and evaluation[5] is not None else "--",
        }

        # Calculate attendance percentage
        cursor.execute("SELECT COUNT(*) FROM main_attendance WHERE student_id = %s AND subject_id = %s", [student_id, subject_id])
        total_attendance = cursor.fetchone()[0] or 0

        cursor.execute("SELECT COUNT(*) FROM main_attendance WHERE student_id = %s AND subject_id = %s AND status = 'present'", [student_id, subject_id])
        present_attendance = cursor.fetchone()[0] or 0

        attendance_percentage = round((present_attendance / total_attendance) * 100, 2) if total_attendance > 0 else "--"

        # Fetch marks percentage
        cursor.execute("SELECT mark_percentage FROM main_marks WHERE student_id = %s AND subject_id = %s", [student_id, subject_id])
        marks_percentage = cursor.fetchone()
        marks_percentage = marks_percentage[0] if marks_percentage and marks_percentage[0] is not None else "--"

        # Calculate quiz performance percentage
        cursor.execute("""
            SELECT COUNT(*) FROM main_quizresponse r
            JOIN main_quizquestions q ON r.question_id = q.id
            JOIN main_quizzes mq ON q.quiz_id = mq.id
            WHERE r.student_id = %s AND mq.subject_id = %s
        """, [student_id, subject_id])
        total_quizzes = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT COUNT(*) FROM main_quizresponse r
            JOIN main_quizquestions q ON r.question_id = q.id
            JOIN main_quizzes mq ON q.quiz_id = mq.id
            WHERE r.student_id = %s AND mq.subject_id = %s AND r.student_response = q.correct_option
        """, [student_id, subject_id])
        correct_quizzes = cursor.fetchone()[0] or 0

        quiz_percentage = round((correct_quizzes / total_quizzes) * 100, 2) if total_quizzes > 0 else "--"

    # Prepare graph data
    graph_data = [
        marks_percentage if marks_percentage != "--" else 0,
        attendance_percentage if attendance_percentage != "--" else 0,
        evaluations["study_time_rating"] if evaluations["study_time_rating"] != "--" else 0,
        evaluations["sleep_time_rating"] if evaluations["sleep_time_rating"] != "--" else 0,
        evaluations["class_participation_rating"] if evaluations["class_participation_rating"] != "--" else 0,
        evaluations["academic_activity_rating"] if evaluations["academic_activity_rating"] != "--" else 0,
        quiz_percentage if quiz_percentage != "--" else 0
    ]

    context = {
        "student": student_data,
        "evaluations": evaluations,
        "attendance_percentage": attendance_percentage,
        "marks_percentage": marks_percentage,
        "quiz_percentage": quiz_percentage,
        "graph_data": json.dumps(graph_data)  # Ensure JSON formatting
    }

    return render(request, 'subject_head/subject_head_students.html', context)




######################################################################################################################

def student_login(request):
    form = StudentLoginForm()

    if request.method == "POST":
        form = StudentLoginForm(request.POST)

        if form.is_valid():
            roll_no = form.cleaned_data.get("roll_no")
            password = form.cleaned_data.get("password")

            # Validate credentials with raw SQL query
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT s.id FROM main_students s
                    JOIN main_users u ON s.user_id = u.id
                    WHERE s.roll_no = %s AND s.password = %s AND u.role = 'student'
                """, [roll_no, password])
                student = cursor.fetchone()

            if student:
                request.session['student_id'] = student[0]  # Save student ID in session
                return redirect('student_dashboard')
            else:
                messages.error(request, 'Invalid roll number or password.')

    return render(request, 'students/student_login.html', {'form': form})


def student_dashboard(request):
    # Ensure the student is logged in
    if 'student_id' not in request.session:
        return redirect('student_login')
    student_id = request.session['student_id']

    print(f"Using student_id: {student_id} for the query.")

    with connection.cursor() as cursor:
        # Fetch the student's name and class ID
        cursor.execute("""
            SELECT s.name, s.class_obj_id
            FROM main_students s 
            WHERE s.id = %s
        """, [student_id])
        student_data = cursor.fetchone()

        if not student_data:
            return redirect('student_login')  # Redirect if student data is invalid

        student_name, class_id = student_data

        class_name = None
        subject_count = 0  # Default to 0 if no subjects are assigned

        if class_id:  # If the student is assigned to a class
            # Fetch the class name
            cursor.execute("""
                SELECT c.class_name 
                FROM main_classes c
                WHERE c.id = %s
            """, [class_id])
            class_data = cursor.fetchone()

            if class_data:
                class_name = class_data[0]

            # Count the number of subjects for this class
            cursor.execute("""
                SELECT COUNT(*) 
                FROM main_subjects 
                WHERE class_obj_id = %s
            """, [class_id])
            subject_count = cursor.fetchone()[0]  # Get the count value

    # Pass data to the template
    return render(request, 'students/student_dashboard.html', {
        'student_name': student_name,
        'class_name': class_name if class_name else "Not Assigned",
        'subject_count': subject_count,
    })


def student_profile(request):
    # Ensure the student is logged in
    if 'student_id' not in request.session:
        return redirect('student_login')

    student_id = request.session['student_id']

    if request.method == 'POST':
        # Handle the form submission and update the student details
        name = request.POST.get('name')
        password = request.POST.get('password')

        with connection.cursor() as cursor:
            # Update student profile in the database
            cursor.execute("""
                UPDATE main_students
                SET name = %s
                WHERE id = %s
            """, [name, student_id])

            if password:  # Only update password if a new one is provided
                cursor.execute("""
                    UPDATE main_students
                    SET password = %s
                    WHERE id = %s
                """, [password, student_id])

        # Redirect after updating the profile
        return redirect('student_profile')

    with connection.cursor() as cursor:
        # Fetch student profile data including current password
        cursor.execute("""
            SELECT s.name, s.roll_no, c.class_name, c.class_head, c.email, i.institution_name, s.password
            FROM main_students s
            LEFT JOIN main_classes c ON s.class_obj_id = c.id
            LEFT JOIN main_institution i ON c.institution_id = i.institution_id
            WHERE s.id = %s
        """, [student_id])
        profile_data = cursor.fetchone()

    # If student data is missing (edge case), redirect to login
    if not profile_data:
        return redirect('student_login')

    student_name, roll_no, class_name, teacher_name, teacher_email, institution_name, current_password = profile_data

    return render(request, 'students/student_profile.html', {
        'student_name': student_name,
        'roll_no': roll_no,
        'class_name': class_name if class_name else "Not Assigned",
        'teacher_name': teacher_name if teacher_name else "Not Assigned",
        'teacher_email': teacher_email if teacher_email else "Not Assigned",
        'institution_name': institution_name if institution_name else "Not Assigned",
        'current_password': current_password,
    })


def student_class(request):
    # Ensure the student is logged in
    if 'student_id' not in request.session:
        return redirect('student_login')

    student_id = request.session['student_id']

    with connection.cursor() as cursor:
        # Fetch student details and class head
        cursor.execute("""
            SELECT s.name AS student_name, c.class_name, c.class_head, c.id AS class_id
            FROM main_students s
            LEFT JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [student_id])
        student_details = cursor.fetchone()

        # Fetch list of students in the same class
        cursor.execute("""
            SELECT s.roll_no, s.name
            FROM main_students s
            WHERE s.class_obj_id = (
                SELECT class_obj_id
                FROM main_students
                WHERE id = %s
            )
        """, [student_id])
        students = cursor.fetchall()

        # Fetch subjects assigned to this class
        cursor.execute("""
            SELECT id, subject_name, subject_head, email
            FROM main_subjects
            WHERE class_obj_id = (
                SELECT class_obj_id
                FROM main_students
                WHERE id = %s
            )
        """, [student_id])
        subjects = cursor.fetchall()

        # Fetch announcements for this class
        cursor.execute("""
            SELECT id, message, created_at
            FROM main_announcements
            WHERE class_obj_id = (
                SELECT class_obj_id
                FROM main_students
                WHERE id = %s
            )
            ORDER BY created_at DESC
        """, [student_id])
        announcements = cursor.fetchall()

    # Handle missing student or class data
    if not student_details:
        student_name, class_name, class_head, class_id = "Unknown", "Not Assigned", "Not Assigned", None
    else:
        student_name, class_name, class_head, class_id = student_details

    students_list = [{'roll_no': row[0], 'name': row[1]} for row in students] if students else []
    subjects_list = [{'id': row[0], 'subject_name': row[1], 'subject_head': row[2], 'email': row[3]} for row in subjects] if subjects else []
    announcements_list = [{'id': row[0], 'message': row[1], 'created_at': row[2]} for row in announcements] if announcements else []

    return render(request, 'students/student_class.html', {
        'student_name': student_name,
        'class_name': class_name,
        'class_head': class_head,
        'students': students_list,
        'student_count': len(students_list),
        'subjects': subjects_list,
        'announcements': announcements_list,  # Pass announcements to the template
    })



def student_chat(request):
    student_id = request.session.get('student_id')  # Get student ID from session
    if not student_id:
        return redirect('student_login')  # Redirect if not logged in

    # Query to get class_head and subject_heads involved in chat with the logged-in student
    query_recent_chats = '''
        SELECT DISTINCT u.id, u.role
        FROM main_chat c
        JOIN main_users u ON (c.sender_id = u.id OR c.receiver_id = u.id)
        JOIN main_students s ON (s.user_id = c.sender_id OR s.user_id = c.receiver_id)
        WHERE s.id = %s
        AND u.role IN ('class_head', 'subject_head')
    '''
    with connection.cursor() as cursor:
        cursor.execute(query_recent_chats, [student_id])
        users_in_chat = cursor.fetchall()  # Returns [(user_id, role), (user_id, role), ...]

    # Get user IDs from the chat query result
    user_ids_in_chat = [user[0] for user in users_in_chat]

    # Fetch class head names and their user IDs for the users involved in recent chats
    class_heads = list(
        Classes.objects.filter(user_id__in=user_ids_in_chat).values_list('user_id', 'class_head')
    )

    # Fetch subject head names and their user IDs for the users involved in recent chats
    subject_heads = list(
        Subjects.objects.filter(user_id__in=user_ids_in_chat).values_list('user_id', 'subject_head')
    )

    # Query to get suggested users (class head of the student's class + subject heads)
    query_suggested_users = '''
        SELECT u.id, u.role, 
               CASE 
                   WHEN u.role = 'class_head' THEN c.class_head
                   WHEN u.role = 'subject_head' THEN s.subject_head
               END AS name
        FROM main_users u
        LEFT JOIN main_classes c ON u.id = c.user_id
        LEFT JOIN main_subjects s ON u.id = s.user_id
        WHERE u.role IN ('class_head', 'subject_head')
        AND ((c.class_head IS NOT NULL AND c.class_head = (SELECT class_head FROM main_classes WHERE id = (SELECT class_obj_id FROM main_students WHERE id = %s)))
             OR s.subject_head IS NOT NULL)
        AND u.id NOT IN (
            SELECT DISTINCT u.id
            FROM main_chat c
            JOIN main_users u ON (c.sender_id = u.id OR c.receiver_id = u.id)
            JOIN main_students st ON (st.user_id = c.sender_id OR st.user_id = c.receiver_id)
            WHERE st.id = %s
        );
    '''
    with connection.cursor() as cursor:
        cursor.execute(query_suggested_users, [student_id, student_id])
        suggested_users = cursor.fetchall()  # Returns [(user_id, role, name), ...]

    return render(request, 'students/student_chat.html', {
        'class_heads': class_heads,  # Contains [(user_id, class_head_name), ...]
        'subject_heads': subject_heads,  # Contains [(user_id, subject_head_name), ...]
        'suggested_users': suggested_users  # Contains [(user_id, role, name), ...]
    })




def student_chat_user(request, user_id):
    student_id = request.session.get('student_id')
    if not student_id:
        return redirect('student_login')

    # Fetch the user_id for the logged-in student
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT s.user_id
            FROM main_students s
            WHERE s.id = %s
        """, [student_id])
        student_user_id = cursor.fetchone()

    if not student_user_id:
        messages.error(request, 'Student not found.')
        return redirect('student_dashboard')

    logged_in_user_id = student_user_id[0]

    # Fetch messages between the logged-in student and the selected user
    query = '''
        SELECT message, sender_id, receiver_id, created_at
        FROM main_chat
        WHERE (sender_id = %s AND receiver_id = %s) OR (sender_id = %s AND receiver_id = %s)
        ORDER BY created_at ASC
    '''
    with connection.cursor() as cursor:
        cursor.execute(query, [logged_in_user_id, user_id, user_id, logged_in_user_id])
        messages_fetched = cursor.fetchall()

    # Get the selected user's role
    user = Users.objects.get(id=user_id)
    selected_user_role = user.role

    # Fetch the selected user's name based on their role
    if selected_user_role == 'class_head':
        selected_user_name = Classes.objects.filter(user_id=user_id).values_list('class_head', flat=True).first()
    elif selected_user_role == 'subject_head':
        selected_user_name = Subjects.objects.filter(user_id=user_id).values_list('subject_head', flat=True).first()
    else:
        selected_user_name = 'Unknown'

    # Handle message submission (POST request)
    if request.method == 'POST':
        message_text = request.POST.get('message')
        if message_text:
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO main_chat (sender_id, receiver_id, message, created_at)
                    VALUES (%s, %s, %s, NOW())
                """, [logged_in_user_id, user_id, message_text])
            return redirect('student_chat_user', user_id=user_id)

    return render(request, 'students/student_chat_user.html', {
        'messages': messages_fetched,
        'user_id': user_id,
        'logged_in_user_id': logged_in_user_id,  # FIX: Pass logged-in user ID for proper message alignment
        'selected_user_name': selected_user_name,
        'selected_user_role': selected_user_role,
    })


def student_subject_detail(request, subject_id):
    if 'student_id' not in request.session:
        return redirect('student_login')

    student_id = request.session['student_id']

    with connection.cursor() as cursor:
        # Fetch subject details along with class information
        cursor.execute("""
            SELECT s.id, s.subject_name, s.subject_head, s.email, c.class_name, c.class_head
            FROM main_subjects s
            JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [subject_id])
        subject_data = cursor.fetchone()

        if not subject_data:
            return render(request, 'students/student_subject_detail.html', {'error': 'Subject not found'})

        # Fetch student name
        cursor.execute("SELECT name FROM main_students WHERE id = %s", [student_id])
        student_name_data = cursor.fetchone()
        student_name = student_name_data[0] if student_name_data else "Unknown"

        # Fetch all quizzes for this subject
        cursor.execute("SELECT id, name FROM main_quizzes WHERE subject_id = %s", [subject_id])
        quizzes = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]

        # Fetch study materials for this subject
        cursor.execute("""
            SELECT id, file_url, created_at, announcement 
            FROM main_studymaterials 
            WHERE subject_id = %s
        """, [subject_id])
        studys = [
            {'id': row[0], 'file_url': row[1], 'created_at': row[2], 'announcement': row[3]}
            for row in cursor.fetchall()
        ]

    # Prepare subject dictionary
    subject = {
        'id': subject_data[0],
        'subject_name': subject_data[1],
        'subject_head': subject_data[2],
        'email': subject_data[3],
        'class_name': subject_data[4],
        'class_head': subject_data[5]
    }

    return render(request, 'students/student_subject_detail.html', {
        'subject': subject,
        'student_name': student_name,
        'quizzes': quizzes,  # Pass quizzes to the template
        'studys': studys  # Pass study materials to the template
    })



def student_quiz(request, subject_id, quiz_id):
    if 'student_id' not in request.session:
        return redirect('student_login')

    student_id = request.session['student_id']
    print(f"Student ID from session: {student_id}")

    # Fetch student name
    with connection.cursor() as cursor:
        cursor.execute("SELECT name FROM main_students WHERE id = %s", [student_id])
        student = cursor.fetchone()
        student_name = student[0] if student else "Unknown"

    print(f"Fetched Student Name: {student_name}")

    # Fetch subject details
    subject = get_object_or_404(Subjects, id=subject_id)
    print(f"Fetched Subject: {subject.subject_name}, Subject Head: {subject.subject_head}")

    # Fetch quiz name
    with connection.cursor() as cursor:
        cursor.execute("SELECT name FROM main_quizzes WHERE id = %s AND subject_id = %s", [quiz_id, subject_id])
        quiz = cursor.fetchone()

        if not quiz:
            print("Quiz not found!")
            return render(request, 'students/student_quiz.html', {'error': 'Quiz not found'})

        quiz_name = quiz[0]
        print(f"Fetched Quiz Name: {quiz_name}")

    # Check if the student has already attempted the quiz
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(*) FROM main_quizresponse 
            WHERE student_id = %s AND question_id IN 
            (SELECT id FROM main_quizquestions WHERE quiz_id = %s)
        """, [student_id, quiz_id])
        quiz_attempted = cursor.fetchone()[0] > 0

    print(f"Quiz Attempted: {quiz_attempted}")

    if quiz_attempted:
        # Fetch quiz questions and student's responses
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT q.id, q.question, q.option_a, q.option_b, q.option_c, q.option_d, 
                       q.correct_option, r.student_response 
                FROM main_quizquestions q
                LEFT JOIN main_quizresponse r ON q.id = r.question_id AND r.student_id = %s
                WHERE q.quiz_id = %s
            """, [student_id, quiz_id])

            quiz_results = [
                {
                    'id': row[0],
                    'question': row[1],
                    'option_a': row[2],
                    'option_b': row[3],
                    'option_c': row[4],
                    'option_d': row[5],
                    'correct_option': row[6],
                    'student_response': row[7],
                    'is_correct': row[6] == row[7]  # Check if the answer is correct
                }
                for row in cursor.fetchall()
            ]

        print(f"Fetched Quiz Results (Responses): {quiz_results}")

        # Calculate score & percentage
        score = sum(1 for q in quiz_results if q['is_correct'])
        total_questions = len(quiz_results)
        percentage = (score / total_questions) * 100 if total_questions > 0 else 0

        return render(request, 'students/student_quiz.html', {
            'student_name': student_name,
            'subject': subject,
            'quiz_id': quiz_id,
            'quiz_name': quiz_name,  # Include quiz name
            'quiz_results': quiz_results,
            'quiz_attempted': True,
            'score': score,
            'total_questions': total_questions,
            'percentage': round(percentage, 2)  # Round to 2 decimal places
        })

    # Fetch quiz questions (only if quiz is not attempted)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, question, option_a, option_b, option_c, option_d 
            FROM main_quizquestions WHERE quiz_id = %s
        """, [quiz_id])
        quiz_questions = [
            {
                'id': row[0],
                'question': row[1],
                'option_a': row[2],
                'option_b': row[3],
                'option_c': row[4],
                'option_d': row[5]
            }
            for row in cursor.fetchall()
        ]

    print(f"Fetched Quiz Questions: {quiz_questions}")

    if request.method == "POST":
        print("Processing Quiz Submission...")

        with connection.cursor() as cursor:
            for key in request.POST:
                if key.startswith("question_"):  # Ensures only quiz responses are processed
                    q_id = key.split("_")[1]  # Extract question ID
                    response = request.POST.get(key)  # Get student's answer

                    print(f"Processing Response -> Question ID: {q_id}, Answer: {response}")

                    if response:  # Insert only if an option is selected
                        cursor.execute("""
                            INSERT INTO main_quizresponse (student_response, question_id, student_id)
                            VALUES (%s, %s, %s)
                        """, [response, q_id, student_id])

                        print(f"Inserted Response -> Question ID: {q_id}, Answer: {response}")

        print("Quiz Submission Completed.")
        return redirect('student_quiz', subject_id=subject_id, quiz_id=quiz_id)  # Reload page to show results

    return render(request, 'students/student_quiz.html', {
        'student_name': student_name,
        'subject': subject,
        'quiz_id': quiz_id,
        'quiz_name': quiz_name,  # Include quiz name
        'quiz_questions': quiz_questions,
        'quiz_attempted': False
    })




def student_performance(request):
    if 'student_id' not in request.session:
        print("[DEBUG] Redirecting to student login (no student_id in session).")
        return redirect('student_login')

    student_id = request.session['student_id']
    subject_id = request.GET.get('subject_id')  # Get the selected subject from query params

    print(f"\n[INFO] Student Performance View Loaded for Student ID: {student_id}")
    if subject_id:
        print(f"[INFO] Selected Subject ID: {subject_id}")

    student_data = {}
    evaluations = {}
    quiz_percentage = 0  # Initialize with 0 instead of "--" to avoid issues in float conversion
    subjects = []
    selected_subject_name = None

    try:
        with connection.cursor() as cursor:
            # Fetch student details along with class info
            cursor.execute("""
                SELECT s.name, s.roll_no, c.class_name, s.email, c.id AS class_id
                FROM main_students s
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s
            """, [student_id])
            row = cursor.fetchone()
            
            if row:
                student_data = {
                    "name": row[0],
                    "roll_no": row[1],
                    "class_name": row[2],
                    "email": row[3],
                    "class_id": row[4]
                }
                print(f"[DEBUG] Student Details: {student_data}")
            else:
                print("[ERROR] Student data not found!")
                return redirect('student_dashboard')

        # Fetch subjects for the student's class
        if student_data.get("class_id"):
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, subject_name FROM main_subjects WHERE class_obj_id = %s
                """, [student_data["class_id"]])
                subjects = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]

            # Get the name of the selected subject
            selected_subject_name = next((s["name"] for s in subjects if str(s["id"]) == subject_id), None)
            
            if selected_subject_name:
                print(f"[DEBUG] Selected Subject: {selected_subject_name} (ID: {subject_id})")
            else:
                print("[WARNING] Selected subject not found in student's class subjects.")

        # Fetch student evaluation details for the selected subject
        if subject_id:
            print(f"[DEBUG] Fetching evaluation data for subject: {selected_subject_name} (ID: {subject_id})...")
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT marks_percentage, attendance_percentage,
                           study_time_rating, sleep_time_rating, 
                           class_participation_rating, academic_activity_rating
                    FROM main_studentevaluation
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    evaluations = {
                        "marks_percentage": round(row[0] or 0, 2),
                        "attendance_percentage": round(row[1] or 0, 2),
                        "study_time_rating": round(row[2] or 0, 2),
                        "sleep_time_rating": round(row[3] or 0, 2),
                        "class_participation_rating": round(row[4] or 0, 2),
                        "academic_activity_rating": round(row[5] or 0, 2),
                    }
                    print(f"[DEBUG] Evaluation Data: {evaluations}")
                else:
                    print(f"[WARNING] No evaluation data found for {selected_subject_name}.")

        # Fetch quiz performance for the selected subject
        if subject_id:
            print(f"[DEBUG] Fetching quiz performance for {selected_subject_name} (ID: {subject_id})...")
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT COUNT(*) AS total_attempted,
                           SUM(CASE WHEN q.correct_option = r.student_response THEN 1 ELSE 0 END) AS correct_answers
                    FROM main_quizresponse r
                    JOIN main_quizquestions q ON r.question_id = q.id
                    JOIN main_quizzes mq ON q.quiz_id = mq.id
                    WHERE r.student_id = %s AND mq.subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    total_attempted, correct_answers = row[0], row[1] or 0
                    quiz_percentage = round((correct_answers / total_attempted) * 100, 2) if total_attempted > 0 else 0
                    print(f"[DEBUG] Quiz Performance: Attempted = {total_attempted}, Correct = {correct_answers}, Percentage = {quiz_percentage}%")
                else:
                    print(f"[WARNING] No quiz data found for {selected_subject_name}.")

        # Prepare data for Graph
        graph_data = [
            float(evaluations.get("marks_percentage", 0)),
            float(evaluations.get("attendance_percentage", 0)),
            float(evaluations.get("study_time_rating", 0)),
            float(evaluations.get("sleep_time_rating", 0)),
            float(evaluations.get("class_participation_rating", 0)),
            float(evaluations.get("academic_activity_rating", 0)),
            float(quiz_percentage)
        ]

    except Exception as e:
        print(f"[ERROR] An error occurred: {e}")
        return redirect('error_page')  # Redirect to an error page if needed

    context = {
        "student": student_data,
        "evaluations": evaluations,
        "quiz_percentage": quiz_percentage,
        "subjects": subjects,
        "selected_subject_id": int(subject_id) if subject_id else None,
        "graph_data": graph_data  # Add graph data to the context
    }

    print("[INFO] Final Context Data Prepared for Rendering.")
    return render(request, "students/student_performance.html", context)








# 🔹 Function to preprocess text: correct spelling, remove extra spaces
def preprocess_text(text):
    text = text.strip().lower()  # Remove extra spaces & lowercase text
    text = str(TextBlob(text).correct())  # Correct spelling
    return text


def student_eduke_bot(request):
    """Handles chatbot page rendering and processing chatbot queries."""

    # Redirect if student is not logged in
    if 'student_id' not in request.session:
        return redirect('student_login')

    student_id = request.session.get("student_id")

    # Fetch student details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT main_students.name, main_students.roll_no, main_classes.class_name
            FROM main_students 
            LEFT JOIN main_classes ON main_students.class_obj_id = main_classes.id
            WHERE main_students.id = %s
        """, [student_id])
        student = cursor.fetchone()

    if not student:
        return JsonResponse({"response": "Student not found. Please log in again."})

    student_details = {
        "name": student[0],
        "roll_no": student[1],
        "class": student[2] if student[2] else "Not assigned"
    }

    if request.method == "GET" and "query" in request.GET:
        query = preprocess_text(request.GET.get("query", ""))  # NLP Preprocessing

        # 🔹 Capabilities
        capability_queries = [
            "what can you do", "how can you help me", "can you help me", "what else can you do",
            "what are your features", "what are your capabilities", "what services do you offer",
            "how can you assist me", "tell me your functions", "what kind of help do you provide",
            "explain your abilities", "what tasks can you perform", "how do you work",
            "are you useful", "why should I use you"
        ]
        capabilities_response = (
            "I can help you with:\n"
            "✔ Checking your attendance records 📅\n"
            "✔ Viewing your marks 📊\n"
            "✔ Viewing your quiz score 📅\n"
            "✔ Providing study tips 📖\n"
            "✔ Suggesting ways to improve in weak subjects 🎯\n"
            "✔ Showing the student leaderboard 🏆\n"
            "✔ Helping with exam preparation 📝\n"
            "✔ Providing time management tips ⏳\n"
            "✔ Boosting your motivation 🚀\n"
            "✔ Reducing stress and improving focus 🌿\n"
            "✔ Teaching memorization & learning techniques 🧠\n"
            "✔ Answering common student queries 💡\n"
            "✔ Increase Productivity and improving focus 🌿\n"
            "Feel free to ask me anything! 😊"
        )
        if any(q in query for q in capability_queries):
            return JsonResponse({"response": (capabilities_response)})

        greeting_responses = {
        "hi": lambda name: f"Hey {name}! Hope you're having a great day. How can I assist you in your studies today? 😊",
        "hello": lambda name: f"Hello {name}! I'm here to help you with your academic journey. What do you need help with?",
        "hey": lambda name: f"Hey {name}! What's up? Need any study tips or motivation? I'm here for you! 🎯",
        "good morning": lambda name: f"Good morning, {name}! Wishing you a productive and successful day ahead. ☀️",
        "good afternoon": lambda name: f"Good afternoon, {name}! Keep up the great work. Need help with anything? 😊",
        "good evening": lambda name: f"Good evening, {name}! Hope you had a great day. Let me know if you need any help! 🌆",
        "good night": lambda name: f"Good night, {name}! Don't forget to rest well and recharge for tomorrow! 🌙",
        "what's up": lambda name: f"Not much, {name}! Just here to help you ace your studies. How can I assist you today? 🎓",
        "how are you": lambda name: f"I'm just a chatbot, {name}, but I'm feeling great because I'm here to help you succeed! 🚀",
        "goodbye": lambda name: f"Goodbye, {name}! Keep learning and stay curious. See you next time! 👋",
        "bye": lambda name: f"Bye, {name}! Have a great day and don't forget to take breaks while studying! 🎯",
        "see you": lambda name: f"See you later, {name}! Remember, learning never stops! 📚",
        "take care": lambda name: f"Take care, {name}! Stay focused, stay positive, and keep striving for success! 🌟",
        "thanks": lambda name: f"You're welcome, {name}! I'm always here to help. Keep up the hard work! 🚀",
        "thank you": lambda name: f"You're most welcome, {name}! If you ever need help again, just ask! 😊"
    }

        # 🔹 Greeting Responses
        for greeting, response in greeting_responses.items():
            if query.startswith(greeting):
                return JsonResponse({"response": response(student_details["name"])})  # ✅ Call function with name

        # 🔹 Identity
        identity_queries = [
            "who are you", "what is your name", "tell me about yourself", "introduce yourself",
            "what are you", "who created you", "who made you", "what do people call you",
            "are you a bot", "are you human", "are you a bot", "tell me about eduke bot",
            "what is eduke bot", "what do you do", "explain yourself"
        ]
        identity_responses = [
            "I'm Eduke Bot, your friendly academic assistant! I help students track their progress, manage attendance, and improve their study habits.",
            "Hey! I'm Eduke Bot, designed to assist you with learning, marks, and study strategies!"
        ]
        if any(q in query for q in identity_queries):
            return JsonResponse({"response": random.choice(identity_responses)})

        # 🔹 Motivation Queries
        motivation_queries = [
            "motivate me", "give me motivation", "inspire me", "i feel demotivated", "i feel down",
            "help me stay motivated", "i lack motivation", "i feel like giving up", "keep me motivated",
            "i need encouragement", "boost my confidence", "i'm losing hope", "help me push forward",
            "i feel stuck", "i'm feeling unmotivated", "i need inspiration", "i feel like i'm failing",
            "how do I stay motivated", "i'm struggling to stay focused", "studying is hard for me"
        ]
        motivation_responses = [
            "You're stronger than you think! Every challenge you face is an opportunity to grow and push your limits. Keep pushing forward, stay consistent, and you’ll see that your efforts will eventually lead to success. 💪",
            "Success comes from small daily improvements. Every step, no matter how small, brings you closer to your goals. Keep going, stay focused on your progress, and you'll achieve things you never thought possible. You’re doing great! 🚀",
            "Mistakes help you learn and grow. They're not failures, but stepping stones on the road to success. Keep trying, because each effort and setback is building your strength and resilience. Success will follow when you persist and believe in the process! 🎯",
            "Believe in yourself! Every expert was once a beginner, and every success story started with a single step. Keep learning, stay curious, and embrace every challenge as a learning opportunity. Your journey is unique, and your growth is inevitable. 📖",
        ]

        if any(q in query for q in motivation_queries):
            return JsonResponse({"response": random.choice(motivation_responses)})

        # 🔹 Study Tips Queries
        study_tips_queries = [
            "give me study tips", "how to study better", "help me study", "best way to study",
            "how to stay focused", "how to study effectively", "how to avoid distractions while studying",
            "how to improve concentration", "how to remember what I study", "how to stop procrastinating",
            "how to manage study time", "how to prepare for exams", "how to study smart", "how to revise properly",
            "how to increase learning speed", "how to make studying fun", "how to avoid stress while studying",
            "how to retain information better", "how to create a study schedule", "best techniques for studying"
        ]
        study_tips_responses = [
            "Consistency is key! Instead of cramming all at once, make a study routine and dedicate a little time each day to review what you’ve learned. This steady approach helps you absorb the material better and ensures you’re not overwhelmed when exams come around. Regular review strengthens long-term retention.",
            "Use the Pomodoro technique: Study for 25 minutes, then take a 5-minute break. This method helps you maintain focus and prevents burnout. The short breaks give your mind time to recharge so that you can approach your studies with fresh energy and concentration each time. Try to stay on task during your study sessions, and reward yourself after each cycle!",
            "Practice with past exams and quizzes to improve memory retention. The more you test yourself on the material, the better you’ll remember it. This active recall technique helps you identify areas where you need improvement and boosts your confidence as you get familiar with the format of potential exam questions.",
            "Stay hydrated and well-rested! A healthy mind learns better, so prioritize sleep and drink plenty of water. Dehydration and lack of sleep can impair cognitive function, making it harder to concentrate and retain information. Ensure you get enough rest and keep water nearby while you study to stay sharp and alert."
        ]

        if any(q in query for q in study_tips_queries):
            return JsonResponse({"response": random.choice(study_tips_responses)})

        # 🔹 Exam Preparation Queries
        exam_prep_queries = [
            "how to prepare for exams", "exam preparation tips", "how to study for exams",
            "best way to revise before an exam", "how to stay calm before an exam",
            "how to score high in exams", "how to study last minute", "best way to study for finals",
            "how to avoid exam stress", "how to manage time during exams", "how to deal with exam fear",
            "how to improve exam performance", "best revision techniques", "how to memorize faster for exams",
            "how to create an exam study plan", "how to stay focused during exams", "how to avoid silly mistakes in exams"
        ]
        exam_prep_responses = [
            "Start preparing early and create a study schedule to cover all topics before the exam. ✅ This will give you plenty of time to review, understand, and practice without feeling rushed. Break down each topic into manageable chunks, and stick to your schedule to stay on track and avoid last-minute cramming.",
            "Practice past papers and quizzes to get familiar with the exam pattern! 📚 Doing so helps you understand the format of the questions, improves your time management, and gives you insight into which topics are frequently tested. It’s one of the best ways to prepare for the actual exam experience.",
            "Take short breaks while studying to keep your mind fresh and focused! 💡 Studies show that short, regular breaks increase productivity and prevent burnout. Use these breaks to stretch, grab a snack, or take a quick walk. This helps you recharge and maintain concentration over longer study sessions.",
            "Revise your notes, use flashcards, and teach someone else to reinforce learning. 🎯 Revising actively helps you remember key concepts better, and teaching others is a powerful method to test your understanding. The more you explain a topic, the more solid your grasp will be, making it easier to recall on exam day."
        ]

        if any(q in query for q in exam_prep_queries):
            return JsonResponse({"response": random.choice(exam_prep_responses)})

        # 🔹 Time Management Queries
        time_management_queries = [
            "how to manage time", "time management tips", "how to balance studies and fun",
            "how to complete homework on time", "how to avoid procrastination",
            "best time management techniques for students", "how to create a study schedule",
            "how to stop wasting time", "how to be more productive", "how to stay disciplined with studies",
            "how to avoid distractions while studying", "how to plan my day effectively",
            "how to manage time for exams and assignments", "how to improve focus and concentration",
            "how to balance school and personal life", "how to avoid last-minute studying"
        ]
        time_management_responses = [
            "Prioritize your tasks and set specific study hours to stay on track! ⏳ Identify the most important and time-sensitive tasks first, and break them down into smaller, manageable steps. Scheduling your study time not only helps you stay organized but also ensures you're using your time efficiently to accomplish what needs to be done.",
            "Use a to-do list or a planner to organize your daily tasks effectively. 📅 Writing down your tasks gives you a clear visual of what needs to be accomplished and helps prevent you from feeling overwhelmed. It’s a great way to stay accountable and motivated as you check off each completed task throughout the day.",
            "Break your tasks into smaller goals and reward yourself after completing each one. 🎯 Breaking big tasks into smaller, more manageable chunks makes them seem less daunting and allows you to make steady progress. After completing each goal, take a short break or treat yourself to something small as a reward to stay motivated and energized.",
            "Avoid distractions like social media while studying. Keep your phone away or use study apps. 📵 Distractions can break your focus and waste valuable time. Try using apps that block distracting sites or set specific time slots for social media breaks. Creating a distraction-free environment helps you maximize your productivity and stay in the zone while studying."
        ]

        if any(q in query for q in time_management_queries):
            return JsonResponse({"response": random.choice(time_management_responses)})

        # 🔹 Stress Management Queries
        stress_management_queries = [
            "how to handle stress", "how to reduce stress", "i feel stressed", "i feel overwhelmed",
            "how to stay calm under pressure", "how to handle exam stress",
            "how to deal with academic pressure", "how to relax after studying", 
            "how to overcome fear of exams", "how to stop overthinking about studies",
            "ways to stay mentally healthy as a student", "how to stay positive during exams",
            "how to manage anxiety before an exam", "how to avoid burnout while studying",
            "how to stay confident before a test", "how to handle school workload stress",
            "how to improve mental well-being as a student"
        ]
        stress_management_responses = [
            "Take deep breaths and remind yourself that you are capable of handling challenges. 🌿 Stress is a natural part of life, but you have the strength to manage it. Taking a few moments to breathe deeply can help calm your nervous system and bring you back to the present. Remember, you've overcome obstacles before, and you can do it again!",
            "Practice meditation or listen to calming music to relax your mind. 🎵 Meditation helps clear your mind and lowers stress by promoting mindfulness. Alternatively, listening to calming music or nature sounds can help reduce anxiety and elevate your mood. Create a peaceful atmosphere where you can unwind and recharge.",
            "Talk to a friend or family member about your worries—it helps to share your feelings. 💬 Opening up to someone you trust can lighten your emotional load and provide a fresh perspective. Talking about what’s stressing you out can also help you process your emotions and feel supported, which can ease the burden of stress.",
            "Stay positive and focus on what you can control instead of worrying about the outcome. 😊 Focusing on what you can manage—like your effort, attitude, and next steps—can help you feel more empowered and less overwhelmed. Trust that you are doing your best and that worrying about things outside your control won't help you succeed."
        ]

        if any(q in query for q in stress_management_queries):
            return JsonResponse({"response": random.choice(stress_management_responses)})

        # 🔹 Productivity Boost Queries
        productivity_queries = [
            "how to be more productive", "how to focus better", "how to avoid distractions",
            "How to improve concentration", "how to get more work done",
            "how to stop procrastinating", "how to build good study habits", 
            "best techniques for deep focus", "how to improve attention span",
            "how to avoid wasting time", "how to stay consistent with studies",
            "how to be more disciplined", "how to create a productive study environment",
            "how to manage study sessions effectively", "how to get into study mode quickly",
            "how to develop a strong work ethic", "how to stay motivated to complete tasks"
        ]
        productivity_responses = [
            "Eliminate distractions and set a timer for focused study sessions! ⏳ Create a distraction-free environment by turning off notifications, keeping your workspace clean, and setting a clear goal for your study session. Using a timer, such as with the Pomodoro technique, can help you stay focused for a set amount of time, boosting your productivity and preventing burnout.",
            "Work on one task at a time instead of multitasking to improve efficiency. 🎯 While multitasking may seem efficient, it often leads to mistakes and wasted time as your focus is divided. By concentrating on one task at a time, you can give it your full attention, complete it faster, and produce higher-quality results.",
            "Use the '2-minute rule': If something takes less than 2 minutes, do it immediately. ✅ Whether it's responding to an email, organizing your desk, or making a quick phone call, completing small tasks right away prevents them from piling up. This simple rule helps you stay on top of your responsibilities and reduces the mental load of unfinished tasks.",
            "Get enough sleep and exercise regularly to keep your brain sharp and active! 🧠 Taking care of your body is crucial for maintaining high levels of productivity. Adequate sleep helps with memory retention and focus, while regular exercise boosts energy and reduces stress, making it easier to stay alert and productive throughout the day."
        ]

        if any(q in query for q in productivity_queries):
            return JsonResponse({"response": random.choice(productivity_responses)})

        # 🔹 Memorization & Learning Techniques Queries
        learning_queries = [
            "how to memorize better", "how to improve memory", "best ways to learn new things",
            "how to retain information", "how to understand complex subjects",
            "how to learn faster", "how to improve comprehension skills",
            "how to remember what I study", "how to develop critical thinking",
            "best techniques for active learning", "how to improve problem-solving skills",
            "how to learn difficult concepts", "how to study smart, not hard",
            "how to apply what I learn", "how to improve analytical skills",
            "how to strengthen cognitive skills", "how to boost brainpower for studying",
            "how to effectively take notes", "best strategies for lifelong learning"
        ]
        learning_responses = [
            "Use mnemonic techniques, like acronyms or visualization, to remember information easily! 🧠 Mnemonics are powerful tools that help you link new information to something you already know, making it easier to recall. Visualizing concepts or creating memorable phrases can significantly boost your memory and make learning more enjoyable.",
            
            "Teach someone else what you've learned—it reinforces your understanding! 🎤 When you explain a topic to another person, you solidify your own knowledge. It forces you to organize and articulate your thoughts, and you may discover areas where you need to strengthen your understanding. Teaching is one of the best ways to master material.",
            
            "Break information into smaller chunks and review it regularly instead of cramming. 📖 Instead of overwhelming yourself with long study sessions, break your material into bite-sized pieces and review them consistently. This method improves long-term retention and makes studying less stressful. The more you review, the easier it becomes to remember.",
            
            "Try using spaced repetition apps to enhance long-term retention! 🔄 Spaced repetition uses algorithms to schedule reviews of material at increasing intervals, which helps you retain information for the long term. Apps like Anki or Quizlet are great for this technique and can boost your recall by ensuring you review key concepts just before you're about to forget them."
        ]

        if any(q in query for q in learning_queries):
            return JsonResponse({"response": random.choice(learning_responses)})
        

        # 🔹 Best & Worst Subjects Query
        if "best and worst subjects" in query or "strongest and weakest subjects" in query:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT main_subjects.subject_name, main_marks.mark_percentage 
                    FROM main_marks 
                    JOIN main_subjects ON main_marks.subject_id = main_subjects.id
                    WHERE main_marks.student_id = %s
                """, [student_id])
                marks_data = cursor.fetchall()

            if marks_data:
                best_subject = max(marks_data, key=lambda x: x[1])
                worst_subject = min(marks_data, key=lambda x: x[1])

                return JsonResponse({
                    "response": f"Your best subject is {best_subject[0]} with {best_subject[1]}% marks. "
                                f"Your weakest subject is {worst_subject[0]} with {worst_subject[1]}%. Keep practicing!"
                })

        # 🔹 Attendance Queries
        if "attendance" in query or "how many classes" in query:
            with connection.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM main_attendance WHERE student_id = %s", [student_id])
                total_classes = cursor.fetchone()[0] or 0

                cursor.execute("SELECT COUNT(*) FROM main_attendance WHERE student_id = %s AND status = 'present'", [student_id])
                present_classes = cursor.fetchone()[0] or 0

            if total_classes > 0:
                attendance_percentage = (present_classes / total_classes) * 100
                motivation = "You're doing great! Keep attending! 😊" if attendance_percentage > 75 else "Try to attend more classes to stay on track. 📚"
                return JsonResponse({"response": f"You attended {present_classes}/{total_classes} classes ({attendance_percentage:.2f}%). {motivation}"})

        # 🔹 Quiz Score Query
        if "quiz" in query and ("score" in query or "quiz performance" in query):
            print("📌 [DEBUG] Quiz score query detected.")  

            with connection.cursor() as cursor:
                sql_query = """
                    SELECT main_quizzes.name, 
                        COUNT(*) AS total_questions, 
                        SUM(CASE WHEN main_quizresponse.student_response = main_quizquestions.correct_option THEN 1 ELSE 0 END) AS correct_answers
                    FROM main_quizresponse
                    JOIN main_quizquestions ON main_quizresponse.question_id = main_quizquestions.id
                    JOIN main_quizzes ON main_quizquestions.quiz_id = main_quizzes.id
                    WHERE main_quizresponse.student_id = %s
                    GROUP BY main_quizzes.name
                    ORDER BY correct_answers DESC
                """

                cursor.execute(sql_query, [student_id])
                quiz_results = cursor.fetchall()

            if quiz_results:
                quiz_text = "\n".join(
                    f"📖 {quiz[0]}: {quiz[2]}/{quiz[1]} correct ({(quiz[2] / quiz[1]) * 100:.2f}%)"
                    for quiz in quiz_results
                )
                return JsonResponse({"response": f"Here are your quiz scores:\n{quiz_text}\nKeep practicing to improve! 🚀"})
            else:
                return JsonResponse({"response": "You haven't taken any quizzes yet. Try one to test your knowledge! 📚"})


        # 🔹 Student Evaluation Query (Study, Sleep, Class Participation, Academic Activity)
        if any(keyword in query for keyword in ["focus", "study", "improve", "stress", "concentration", "sleep", "class participation", "academic activity"]):
            print("📌 [DEBUG] Student evaluation query detected.")

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT study_time_rating, sleep_time_rating, class_participation_rating, academic_activity_rating
                    FROM main_studentevaluation WHERE student_id = %s
                """, [student_id])
                evaluations = cursor.fetchall()

            if evaluations:
                # Initialize sum and count for calculating averages
                total_study_time, total_sleep_time, total_class_participation, total_academic_activity = 0, 0, 0, 0
                count_study_time, count_sleep_time, count_class_participation, count_academic_activity = 0, 0, 0, 0

                # Process each evaluation row
                for study_time, sleep_time, class_participation, academic_activity in evaluations:
                    if study_time is not None:
                        total_study_time += study_time
                        count_study_time += 1
                    if sleep_time is not None:
                        total_sleep_time += sleep_time
                        count_sleep_time += 1
                    if class_participation is not None:
                        total_class_participation += class_participation
                        count_class_participation += 1
                    if academic_activity is not None:
                        total_academic_activity += academic_activity
                        count_academic_activity += 1

                # Calculate averages (handling NULL values)
                avg_study_time = total_study_time / count_study_time if count_study_time > 0 else None
                avg_sleep_time = total_sleep_time / count_sleep_time if count_sleep_time > 0 else None
                avg_class_participation = total_class_participation / count_class_participation if count_class_participation > 0 else None
                avg_academic_activity = total_academic_activity / count_academic_activity if count_academic_activity > 0 else None

                response = None  # Default response

                # Check query intent and respond specifically
                if "study" in query or "focus" in query or "improve" in query:
                    if avg_study_time is not None:
                        if avg_study_time < 40:
                            response = ("Your study time is quite low, which might be affecting your academic performance. 📚 "
                                        "It's important to establish a structured study routine. Try setting small, achievable goals for each session, "
                                        "and use techniques like the Pomodoro method to stay focused. Also, consider reducing distractions by keeping your study space organized "
                                        "and free from interruptions. Over time, consistency will help you improve significantly! 🚀")
                        elif avg_study_time < 70:
                            response = ("You're doing a decent job with your studies, but there's room for improvement. 💡 "
                                        "Try to analyze which subjects or topics need more focus and allocate additional time to them. "
                                        "Studying smarter is just as important as studying harder—use active recall, note-making, and mind maps to make learning more effective. "
                                        "Setting deadlines and rewarding yourself after completing study goals can keep you motivated. Keep pushing forward!")
                        else:
                            response = ("Fantastic! You're maintaining a great study routine, and it's clearly benefiting your learning. 🚀 "
                                        "To maximize your efforts, consider teaching what you've learned to others—it strengthens your understanding. "
                                        "Also, make sure to take short breaks to avoid burnout and refresh your mind. Keep challenging yourself with new learning strategies, "
                                        "and don't forget to stay curious and enjoy the process of learning! 🎓")
                    else:
                        response = ("I don't have enough study data yet to give you feedback. 📊 "
                                    "However, if you're looking to improve your study habits, start by tracking your daily study hours and the topics you cover. "
                                    "Small, steady improvements will make a big difference over time!")

                elif "sleep" in query:
                    if avg_sleep_time is not None:
                        if avg_sleep_time < 40:
                            response = ("You're not getting enough sleep, and this could be affecting your concentration, memory, and overall health. 💤 "
                                        "Lack of sleep can lead to decreased focus, slower cognitive function, and even increased stress levels. "
                                        "Try to set a fixed bedtime and wake-up schedule, avoid screen time before sleeping, and create a relaxing bedtime routine. "
                                        "Aim for at least 7-8 hours of sleep every night, as quality rest is essential for learning and brain function. 🌙")
                        elif avg_sleep_time < 70:
                            response = ("You're getting a moderate amount of sleep, but improving your sleep quality could enhance your energy levels and focus. 🌙 "
                                        "Consider maintaining a regular sleep schedule, avoiding caffeine before bedtime, and ensuring that your sleeping environment is comfortable. "
                                        "Your brain consolidates information while you sleep, so getting better rest can actually help you retain what you study more effectively!")
                        else:
                            response = ("You're doing an excellent job maintaining a good sleep schedule! ✅ "
                                        "Quality sleep is one of the key factors in academic success. It helps with focus, mood regulation, and energy levels. "
                                        "Just make sure you continue this healthy habit, and if you ever feel fatigued despite enough sleep, consider your diet and exercise routine as well. "
                                        "Keep up the great work, and your brain will thank you!")
                    else:
                        response = ("I don't have enough sleep data yet, but if you're looking to improve your sleep quality, start tracking your sleep patterns. 🌙 "
                                    "A well-rested mind is more productive and efficient, so prioritize good sleep habits!")

                elif "class participation" in query:
                    if avg_class_participation is not None:
                        if avg_class_participation < 40:
                            response = ("It looks like you aren't participating much in class discussions. 🎤 "
                                        "Active participation can help you understand topics better and improve your confidence. "
                                        "Try asking questions, sharing your thoughts, or even taking small steps like nodding and making eye contact with the teacher. "
                                        "Engaging with the class material and your peers can make learning more interactive and enjoyable!")
                        elif avg_class_participation < 70:
                            response = ("You're participating fairly well in class, but there's still room to be more involved. ✨ "
                                        "Try contributing more by answering questions, discussing ideas with classmates, and sharing your opinions on topics. "
                                        "Even small efforts like summarizing what the teacher said in your own words can boost understanding and retention. "
                                        "Active learning is a great way to build confidence and sharpen your thinking skills!")
                        else:
                            response = ("You're doing a fantastic job actively participating in class! 📖 "
                                        "Engaging in discussions and asking questions not only helps you learn better but also makes the learning process more enjoyable. "
                                        "Keep up the enthusiasm and continue being involved—your curiosity and effort will take you far!")
                    else:
                        response = ("I don't have enough class participation data yet, but remember that speaking up and engaging in discussions can improve your understanding. 🎤 "
                                    "Start by answering simple questions, sharing your thoughts, and gradually building the confidence to express yourself more in class!")

                elif "academic activity" in query:
                    if avg_academic_activity is not None:
                        if avg_academic_activity < 40:
                            response = ("It seems like you're not engaging much in academic activities outside regular studies. 📊 "
                                        "Taking part in extra learning opportunities, such as solving additional exercises, joining study groups, or participating in academic clubs, "
                                        "can help you develop a deeper understanding of subjects. Try exploring online courses, competitions, or group discussions—they can make learning fun and effective!")
                        elif avg_academic_activity < 70:
                            response = ("You're doing a decent job engaging in academic activities, but pushing yourself a little further can be beneficial. 📌 "
                                        "Consider exploring new learning resources like educational videos, books, and online practice tests. "
                                        "Participating in academic discussions or challenging yourself with new problems can help you develop a broader perspective. "
                                        "Keep up the momentum, and soon, you'll see improvements in your critical thinking and analytical skills!")
                        else:
                            response = ("You're actively engaged in academic activities, and that's a fantastic habit! 🎯 "
                                        "Your curiosity and willingness to explore beyond textbooks are what set high achievers apart. "
                                        "Keep challenging yourself with new learning experiences, and don't hesitate to take up leadership roles in academic projects or mentor your peers. "
                                        "Your dedication will lead to great success!")
                    else:
                        response = ("I don't have enough academic activity data yet, but participating in extra learning activities can make a big difference. 📚 "
                                    "Try joining academic clubs, taking part in discussions, or engaging in self-learning outside class—these efforts can greatly enhance your knowledge!")

                else:
                    response = ("I'm not sure how to interpret your query. Could you clarify what you're looking for? 😊")


                return JsonResponse({"response": response})
            else:
                return JsonResponse({"response": "I don't have any evaluation data yet. Keep up the good work and check back later! 📊"})


        # 🔹 Leaderboard Query
        if "leaderboard" in query or "top students" in query:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT name, AVG(mark_percentage) as avg_marks FROM main_students 
                    JOIN main_marks ON main_students.id = main_marks.student_id
                    GROUP BY main_students.id ORDER BY avg_marks DESC LIMIT 5
                """)
                leaderboard = cursor.fetchall()

            leaderboard_text = "\n".join(f"{i+1}. {row[0]} - {row[1]:.2f}%" for i, row in enumerate(leaderboard))
            return JsonResponse({"response": f"🏆 Top Performing Students:\n{leaderboard_text}\n\nKeep working hard! 💪"})


        # 🔹 Default Response
        return JsonResponse({"response": "I'm here to help with attendance, marks, and study tips! Just ask! 😊"})

    return render(request, "students/student_eduke_bot.html", {"student": student_details})



######################################################################################################################

def parent_login(request):
    print("🔹 Parent login view called.")

    form = ParentLoginForm()

    if request.method == "POST":
        print("🔹 POST request received.")

        form = ParentLoginForm(request.POST)

        if form.is_valid():
            print("✅ Form is valid.")
            roll_no = form.cleaned_data.get("roll_no")
            password = form.cleaned_data.get("password")
            print(f"🔹 Received Roll No: {roll_no}, Password: {password}")

            # Validate credentials with raw SQL query
            with connection.cursor() as cursor:
                print("🔹 Executing SQL query to authenticate parent...")
                cursor.execute("""
                    SELECT p.id FROM main_parents p
                    JOIN main_students s ON p.student_id = s.id
                    JOIN main_users u ON p.user_id = u.id
                    WHERE s.roll_no = %s AND p.password = %s AND u.role = 'parent'
                """, [roll_no, password])
                parent = cursor.fetchone()
                print(f"🔹 Query Result: {parent}")

            if parent:
                # Save the parent ID in session
                request.session['parent_id'] = parent[0]
                print(f"✅ Session Created: parent_id = {request.session.get('parent_id')}")
                messages.success(request, "Login successful.")
                print("🔹 Redirecting to parent_dashboard...")
                return redirect('parent_dashboard')  # Redirect to parent dashboard
            else:
                print("❌ Invalid credentials.")
                messages.error(request, "Invalid roll number or password.")

    print("🔹 Rendering login page again.")
    return render(request, 'parents/parent_login.html', {'form': form})
    


def parent_dashboard(request):
    print("🔹 Parent Dashboard View Called")

    # Debug session data
    print(f"🔹 Session Data: {dict(request.session.items())}")

    parent_id = request.session.get('parent_id')
    if not parent_id:
        print("❌ No parent_id in session. Redirecting to login.")
        return redirect('parent_login')

    with connection.cursor() as cursor:
        print(f"🔹 Fetching parent details for parent_id: {parent_id}")
        
        cursor.execute("""
            SELECT p.id, p.name AS parent_name, p.student_id, 
                   s.id AS student_id, s.name AS student_name, s.class_obj_id
            FROM main_parents p
            JOIN main_students s ON p.student_id = s.id  
            WHERE p.id = %s
        """, [parent_id])
        
        parent_data = cursor.fetchone()
        
        if not parent_data:
            print(f"❌ No parent found with id {parent_id}. Redirecting to login.")
            return redirect('parent_login')

        print(f"✅ Parent Data Fetched: {parent_data}")

        # Correctly unpacking all 6 values
        parent_id, parent_name, student_roll_no, student_id, student_name, student_class_id = parent_data

        print(f"🔹 Fetching class details for class_id: {student_class_id}")
        cursor.execute("""
            SELECT class_name, class_head FROM main_classes
            WHERE id = %s
        """, [student_class_id])
        class_data = cursor.fetchone()

        if class_data:
            student_class, class_teacher = class_data
            print(f"✅ Class Data: {class_data}")
        else:
            student_class, class_teacher = "Class not assigned", "No teacher assigned"
            print("❌ Class data not found.")

        # Fetch subjects count
        print(f"🔹 Fetching subject count for class_id: {student_class_id}")
        cursor.execute("""
            SELECT COUNT(*) FROM main_subjects
            WHERE class_obj_id = %s
        """, [student_class_id])
        subjects_count_data = cursor.fetchone()
        subjects_count = subjects_count_data[0] if subjects_count_data else 0
        print(f"✅ Subjects Count: {subjects_count}")

    # Prepare context for the template
    context = {
        'parent_id': parent_id,
        'parent_name': parent_name,
        'student_roll_no': student_roll_no,  # Added
        'student_id': student_id,  # Added
        'student_name': student_name,
        'student_class': student_class,
        'class_teacher': class_teacher,
        'subjects_count': subjects_count
    }

    print("✅ Rendering Parent Dashboard Page")
    return render(request, 'parents/parent_dashboard.html', context)



def parent_teacher_profile(request):
    parent_id = request.session.get('parent_id')  # Ensure the parent is logged in
    if not parent_id:
        return redirect('parent_login')

    # Fetch class teacher details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT c.class_head AS teacher_name, c.email AS teacher_email, 
                   c.class_name, i.institution_name
            FROM main_parents p
            JOIN main_students s ON p.student_id = s.id
            LEFT JOIN main_classes c ON s.class_obj_id = c.id
            LEFT JOIN main_institution i ON c.institution_id = i.institution_id
            WHERE p.id = %s
        """, [parent_id])
        teacher_data = cursor.fetchone()

    # Handle case where teacher data is not found
    if not teacher_data:
        messages.error(request, "Unable to retrieve teacher details.")
        return redirect('parent_dashboard')

    # Prepare context data for the template
    context = {
        'teacher_name': teacher_data[0] or "N/A",
        'teacher_email': teacher_data[1] or "N/A",
        'class_name': teacher_data[2] or "N/A",
        'institution_name': teacher_data[3] or "N/A",
    }
    return render(request, 'parents/parent_teacher_profile.html', context)



def parent_class(request):
    print("🔹 Parent Class View Called")

    parent_id = request.session.get('parent_id')
    print(f"🔹 Session Data: {request.session.items()}")  # Debug session

    if not parent_id:
        print("❌ No parent session found. Redirecting to login.")
        return redirect('parent_login')

    with connection.cursor() as cursor:
        # Fetch parent and student details
        cursor.execute("""
            SELECT p.id, p.name AS parent_name, p.student_id, 
                   s.name AS student_name, s.class_obj_id 
            FROM main_parents p
            JOIN main_students s ON p.student_id = s.id
            WHERE p.id = %s
        """, [parent_id])
        parent_data = cursor.fetchone()
        print(f"🔹 Parent Data Fetched: {parent_data}")

        if not parent_data:
            print("❌ Parent data not found. Redirecting to login.")
            return redirect('parent_login')

        # Extract details
        _, parent_name, student_roll_no, student_name, student_class_id = parent_data

        # Fetch class details
        cursor.execute("""
            SELECT class_name, class_head FROM main_classes
            WHERE id = %s
        """, [student_class_id])
        class_data = cursor.fetchone()
        student_class = class_data[0] if class_data else "Class not assigned"
        class_teacher = class_data[1] if class_data else "No teacher assigned"
        print(f"🔹 Class Data: {class_data}")

        # Fetch subjects
        cursor.execute("""
            SELECT subject_name, subject_head, email FROM main_subjects
            WHERE class_obj_id = %s
        """, [student_class_id])
        subjects = cursor.fetchall()
        print(f"🔹 Subjects Found: {len(subjects)}")

        # Fetch students
        cursor.execute("""
            SELECT roll_no, name, email FROM main_students
            WHERE class_obj_id = %s
        """, [student_class_id])
        students = cursor.fetchall()
        print(f"🔹 Students Found: {len(students)}")

    # Prepare context for the template
    context = {
        'parent_name': parent_name,
        'student_name': student_name,
        'student_class': student_class,
        'class_teacher': class_teacher,
        'subjects': subjects,
        'students': students,
        'student_count': len(students)
    }

    print("✅ Rendering parent_class.html")
    return render(request, 'parents/parent_class.html', context)



def parent_profile(request):
    parent_id = request.session.get('parent_id')  # Ensure the parent is logged in
    if not parent_id:
        return redirect('parent_login')

    if request.method == 'POST':
        # Handle profile update
        parent_name = request.POST.get('parent_name')
        password = request.POST.get('password')

        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE main_parents 
                SET name = %s, password = %s 
                WHERE id = %s
            """, [parent_name, password, parent_id])
            messages.success(request, "Profile updated successfully!")
        return redirect('parent_profile')

    # Fetch parent's profile details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT p.name AS parent_name, p.password, 
                   s.name AS student_name, s.roll_no, 
                   c.class_name, 
                   c.class_head AS teacher_name, c.email AS teacher_email
            FROM main_parents p
            JOIN main_students s ON p.student_id = s.id
            LEFT JOIN main_classes c ON s.class_obj_id = c.id
            WHERE p.id = %s
        """, [parent_id])
        parent_data = cursor.fetchone()

    # Handle case where profile data is not found
    if not parent_data:
        messages.error(request, "Unable to retrieve profile details.")
        return redirect('parent_dashboard')

    # Prepare context data for the template
    context = {
        'parent_name': parent_data[0] or "",
        'password': parent_data[1] or "",
        'student_name': parent_data[2] or "N/A",
        'student_roll_no': parent_data[3] or "N/A",
        'class_name': parent_data[4] or "N/A",
        'teacher_name': parent_data[5] or "N/A",
        'teacher_email': parent_data[6] or "N/A",
    }
    return render(request, 'parents/parent_profile.html', context)



def parent_chat(request):
    parent_id = request.session.get('parent_id')  # Get parent ID from session
    if not parent_id:
        return redirect('parent_login')  # Redirect to login if not logged in

    print(f"Parent ID for chat: {parent_id}")

    with connection.cursor() as cursor:
        # Get the parent's user ID
        cursor.execute("SELECT user_id FROM main_parents WHERE id = %s", [parent_id])
        user_data = cursor.fetchone()

        if not user_data:
            print("No user_id found for this parent.")
            return render(request, 'parents/parent_chat.html', {'chat_users': [], 'suggested_users': []})

        user_id = user_data[0]
        print(f"User ID for parent: {user_id}")

        # Get recent chat users (Users the parent has already chatted with)
        cursor.execute("""
            SELECT DISTINCT u.id, 
                            COALESCE(c.class_head, s.subject_head) AS name, 
                            u.role
            FROM main_chat ch
            JOIN main_users u ON (ch.sender_id = u.id OR ch.receiver_id = u.id)
            LEFT JOIN main_classes c ON u.id = c.user_id
            LEFT JOIN main_subjects s ON u.id = s.user_id
            WHERE (ch.sender_id = %s OR ch.receiver_id = %s)
            AND u.role IN ('class_head', 'subject_head')
        """, [user_id, user_id])

        chat_users = cursor.fetchall()
        print(f"Recent chat users: {chat_users}")

        recent_chat_user_ids = {user[0] for user in chat_users}  # Extract user IDs
        print(f"Recent chat user IDs: {recent_chat_user_ids}")

        # Get the class_head of the student's class
        cursor.execute("""
            SELECT c.user_id, c.class_head
            FROM main_classes c
            JOIN main_students st ON st.class_obj_id = c.id
            WHERE st.id = (SELECT student_id FROM main_parents WHERE id = %s)
        """, [parent_id])
        class_head = cursor.fetchone()

        class_head_id = class_head[0] if class_head else None
        print(f"Class Head: {class_head}")

        # Get suggested users excluding those in recent chats
        cursor.execute("""
            SELECT u.id, u.role, 
                CASE 
                    WHEN u.role = 'class_head' THEN c.class_head
                    WHEN u.role = 'subject_head' THEN s.subject_head
                END AS name
            FROM main_users u
            LEFT JOIN main_classes c ON u.id = c.user_id
            LEFT JOIN main_subjects s ON u.id = s.user_id
            WHERE u.role IN ('class_head', 'subject_head')
            AND (
                (c.class_head IS NOT NULL AND c.class_head = (SELECT class_head FROM main_classes WHERE id = (SELECT class_obj_id FROM main_students WHERE roll_no = (SELECT student_id FROM main_parents WHERE id = %s)) ))
                OR s.subject_head IS NOT NULL
            )
            AND u.id NOT IN (
                SELECT DISTINCT u.id
                FROM main_chat ch
                JOIN main_users u ON (ch.sender_id = u.id OR ch.receiver_id = u.id)
                WHERE (ch.sender_id = %s OR ch.receiver_id = %s)
            )
        """, [parent_id, user_id, user_id])

        suggested_users = cursor.fetchall()
        suggested_users = list(suggested_users)  # Convert tuple to list to allow modification
        print(f"Suggested users before adding class head: {suggested_users}")

        suggested_user_ids = {user[0] for user in suggested_users}  # Extract user IDs

        # Ensure the class head is suggested if they haven't been messaged yet
        if class_head_id and class_head_id not in recent_chat_user_ids and class_head_id not in suggested_user_ids:
            print(f"Adding class head to suggested users: {class_head}")
            suggested_users.append((class_head_id, 'class_head', class_head[1]))

        print(f"Final suggested users: {suggested_users}")

    return render(request, 'parents/parent_chat.html', {
        'chat_users': chat_users,
        'suggested_users': suggested_users,
    })



def parent_chat_user(request, user_id):
    parent_id = request.session.get('parent_id')
    if not parent_id:
        return redirect('parent_login')

    # Fetch the parent’s user_id
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT p.user_id
            FROM main_parents p
            WHERE p.id = %s
        """, [parent_id])
        parent_user_id = cursor.fetchone()

    if not parent_user_id:
        messages.error(request, 'Parent not found.')
        return redirect('parent_dashboard')

    # Debugging: Print the parent's user_id
    print(f"Parent's user_id for chat: {parent_user_id[0]}")

    # Fetch messages between the logged-in parent and the selected user
    query = '''
        SELECT message, sender_id, receiver_id, created_at
        FROM main_chat
        WHERE (sender_id = %s AND receiver_id = %s) OR (sender_id = %s AND receiver_id = %s)
        ORDER BY created_at ASC
    '''
    with connection.cursor() as cursor:
        cursor.execute(query, [parent_user_id[0], user_id, user_id, parent_user_id[0]])
        messages_fetched = cursor.fetchall()

    # Debugging: Print the fetched messages
    print(f"Messages fetched: {messages_fetched}")

    # Get the selected user's role
    user = Users.objects.get(id=user_id)
    selected_user_role = user.role

    # Fetch the selected user's name based on their role
    if selected_user_role == 'class_head':
        selected_user_name = Classes.objects.filter(user_id=user_id).values_list('class_head', flat=True).first()
    elif selected_user_role == 'subject_head':
        selected_user_name = Subjects.objects.filter(user_id=user_id).values_list('subject_head', flat=True).first()
    else:
        selected_user_name = 'Unknown'

    # Fetch the teacher's user_id for message sender identification
    if selected_user_role == 'class_head':
        teacher_user_id = Classes.objects.filter(class_head=user_id).values_list('user_id', flat=True).first()
    else:
        teacher_user_id = Subjects.objects.filter(subject_head=user_id).values_list('user_id', flat=True).first()

    # Handle message submission (POST request)
    if request.method == 'POST':
        message_text = request.POST.get('message')
        if message_text:
            # Save the new message to the database
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO main_chat (sender_id, receiver_id, message, created_at)
                    VALUES (%s, %s, %s, NOW())
                """, [parent_user_id[0], user_id, message_text])

            # Redirect the user back to the chat page to see the new message
            return redirect('parent_chat_user', user_id=user_id)

    return render(request, 'parents/parent_chat_user.html', {
        'messages': messages_fetched, 
        'user_id': user_id,
        'logged_in_user_id': parent_user_id[0],  # Pass this for sender alignment
        'selected_user_name': selected_user_name,
        'selected_user_role': selected_user_role,
        'teacher_user_id': teacher_user_id
    })


def parent_evaluation(request):
    parent_id = request.session.get('parent_id')  # Get parent ID from session
    if not parent_id:
        print("❌ Parent ID not found in session. Redirecting to login.")
        return redirect('parent_login')

    # Fetch the parent's associated student details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT main_students.id, main_students.name, main_students.roll_no, main_students.class_obj_id
            FROM main_students 
            JOIN main_parents ON main_students.id = main_parents.student_id
            WHERE main_parents.id = %s
        """, [parent_id])
        student = cursor.fetchone()

    if not student:
        print("❌ No student found for this parent.")
        return render(request, 'parents/parent_evaluation.html', {'student': None, 'subjects': []})

    student_id, student_name, roll_no, class_id = student

    # Fetch subjects for the student's class
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, subject_name FROM main_subjects WHERE class_obj_id = %s
        """, [class_id])
        subjects = cursor.fetchall()

    # Fetch existing evaluation data per subject
    evaluation_data = []
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT subject_id, study_time_rating, sleep_time_rating 
            FROM main_studentevaluation 
            WHERE student_id = %s
        """, [student_id])
        evaluations = {row[0]: row for row in cursor.fetchall()}  # Convert to dictionary with subject_id as key

    # Prepare structured evaluation data
    structured_subjects = []
    for subject in subjects:
        subject_id, subject_name = subject
        structured_subjects.append({
            'id': subject_id,
            'name': subject_name,
            'study_time_rating': evaluations.get(subject_id, [None, 0, 0])[1],  # Default 0 if not found
            'sleep_time_rating': evaluations.get(subject_id, [None, 0, 0])[2],  # Default 0 if not found
        })

    # Handle form submission
    if request.method == "POST":
        print("✅ Form submission received.")

        for subject in structured_subjects:
            subject_id = subject['id']
            study_time_rating = request.POST.get(f"study_time_rating_{subject_id}")
            sleep_time_rating = request.POST.get(f"sleep_time_rating_{subject_id}")

            try:
                study_time_rating = float(study_time_rating) if study_time_rating else 0
                sleep_time_rating = float(sleep_time_rating) if sleep_time_rating else 0
            except ValueError:
                print(f"❌ Invalid rating values for subject {subject['name']}. Skipping update.")
                continue

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id FROM main_studentevaluation 
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                existing_record = cursor.fetchone()

                if existing_record:
                    cursor.execute("""
                        UPDATE main_studentevaluation 
                        SET study_time_rating = %s, sleep_time_rating = %s 
                        WHERE student_id = %s AND subject_id = %s
                    """, [study_time_rating, sleep_time_rating, student_id, subject_id])
                else:
                    cursor.execute("""
                        INSERT INTO main_studentevaluation (student_id, subject_id, study_time_rating, sleep_time_rating)
                        VALUES (%s, %s, %s, %s)
                    """, [student_id, subject_id, study_time_rating, sleep_time_rating])

        messages.success(request, "Student evaluation updated successfully!")
        return redirect('parent_evaluation')

    return render(request, 'parents/parent_evaluation.html', {
        'student': {'id': student_id, 'name': student_name, 'roll_no': roll_no},
        'subjects': structured_subjects,  # Now contains the necessary data
    })





from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, redirect
from django.db import connection

def parent_student_performance(request):
    print("🔹 Starting parent_student_performance view...")

    parent_id = request.session.get('parent_id')  # Get parent ID from session
    if not parent_id:
        print("❌ Parent ID not found in session. Redirecting to login.")
        return redirect('parent_login')

    student_data = {}
    evaluations = {}
    quiz_percentage = 0  # Default to 0 instead of "--" for easier float conversion
    subjects = []
    selected_subject_name = None

    subject_id = request.GET.get('subject_id')  # Get selected subject from query params

    print(f"ℹ️ Parent ID from session: {parent_id}")

    try:
        # Fetch student linked to this parent
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT s.id, s.name, s.roll_no, c.class_name, s.email, c.id AS class_id
                FROM main_parents p
                JOIN main_students s ON p.student_id = s.id
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE p.id = %s
            """, [parent_id])
            row = cursor.fetchone()

            if row:
                student_id = row[0]
                student_data = {
                    "name": row[1],
                    "roll_no": row[2],
                    "class_name": row[3],
                    "email": row[4],
                    "class_id": row[5],
                }
                print(f"✅ Student Data Retrieved: {student_data}")
            else:
                print("⚠️ No student linked to this parent.")
                return render(request, "parents/parent_student_performance.html", {"error": "No student data available."})

        # Fetch subjects for the student's class
        if student_data.get("class_id"):
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, subject_name FROM main_subjects WHERE class_obj_id = %s
                """, [student_data["class_id"]])
                subjects = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]

            # Get the name of the selected subject
            selected_subject_name = next((s["name"] for s in subjects if str(s["id"]) == subject_id), None)
            
            if selected_subject_name:
                print(f"🔹 Selected Subject: {selected_subject_name} (ID: {subject_id})")
            else:
                print("⚠️ Selected subject not found in student's class subjects.")

        # Fetch Student Evaluations for the selected subject
        if subject_id:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT marks_percentage, attendance_percentage,
                           study_time_rating, sleep_time_rating, class_participation_rating, academic_activity_rating
                    FROM main_studentevaluation
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    evaluations = {
                        "marks_percentage": round(float(row[0] or 0), 2),
                        "attendance_percentage": round(float(row[1] or 0), 2),
                        "study_time_rating": round(float(row[2] or 0), 2),
                        "sleep_time_rating": round(float(row[3] or 0), 2),
                        "class_participation_rating": round(float(row[4] or 0), 2),
                        "academic_activity_rating": round(float(row[5] or 0), 2),
                    }
                    print(f"📊 Evaluations Retrieved: {evaluations}")
                else:
                    print(f"⚠️ No evaluation data found for {selected_subject_name}.")

        # Fetch Quiz Performance for the selected subject
        if subject_id:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT COUNT(*) AS total_attempted,
                           SUM(CASE WHEN q.correct_option = r.student_response THEN 1 ELSE 0 END) AS correct_answers
                    FROM main_quizresponse r
                    JOIN main_quizquestions q ON r.question_id = q.id
                    JOIN main_quizzes mq ON q.quiz_id = mq.id
                    WHERE r.student_id = %s AND mq.subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    total_attempted, correct_answers = row[0], row[1] or 0
                    quiz_percentage = round((correct_answers / total_attempted) * 100, 2) if total_attempted > 0 else 0
                    print(f"📝 Quiz Performance: Attempted = {total_attempted}, Correct = {correct_answers}, Percentage = {quiz_percentage}%")
                else:
                    print(f"⚠️ No quiz data found for {selected_subject_name}.")

        # Prepare data for Graph
        graph_data = [
            float(evaluations.get("marks_percentage", 0)),
            float(evaluations.get("attendance_percentage", 0)),
            float(evaluations.get("study_time_rating", 0)),
            float(evaluations.get("sleep_time_rating", 0)),
            float(evaluations.get("class_participation_rating", 0)),
            float(evaluations.get("academic_activity_rating", 0)),
            float(quiz_percentage)
        ]

    except Exception as e:
        print(f"❌ An error occurred: {e}")
        return redirect('error_page')

    context = {
        "student": student_data,
        "evaluations": evaluations,
        "quiz_percentage": quiz_percentage,
        "subjects": subjects,
        "selected_subject_id": int(subject_id) if subject_id else None,
        "graph_data": graph_data,  # Pass graph data for chart rendering
    }

    print("✅ Rendering 'parent_student_performance.html' template.")
    return render(request, "parents/parent_student_performance.html", context)



######################################################################################################################


